"""
excel_handler.py
----------------
Le qualquer planilha (.xlsx/.xls/.csv) e gera uma saida enriquecida.

A detecção de cabeçalhos é feita dinamicamente por IA (Mistral) ou por
palavras-chave — o arquivo não precisa mais seguir o formato fixo do Mais Obras.

A planilha enriquecida gerada por este módulo sempre sai assim:
  Linha 1: cabeçalho (colunas originais + colunas de contato)
  Linha 2+: dados
"""

import io
import logging
import re
import unicodedata
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ai_header_detector import detectar_estrutura_planilha, extrair_uf_de_texto

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constantes de formatação
# ---------------------------------------------------------------------------

COLUNAS_NOVAS = [
    "Telefone Arquiteto 1",
    "Telefone Arquiteto 2",
    "Email Arquiteto",
    "Telefone Proprietario 1",
    "Telefone Proprietario 2",
    "Email Proprietario",
    "Status",
]

COR_HEADER = "56181B"
COR_HEADER_NOVO = "CBB068"
COR_HEADER_FONT = "FFF7E2"
COR_HEADER_NOVO_FONT = "56181B"
COR_ZEBRA = "FBF7EF"
COR_BORDA = "E2D7C8"


# ---------------------------------------------------------------------------
# Modelo de dados
# ---------------------------------------------------------------------------

@dataclass
class ObraRow:
    """Representa uma linha do Excel com dados suficientes para buscar no Mais Obras."""

    row_index: int
    nome_profissional: str
    nome_proprietario: str
    cidade: str
    uf: str
    endereco: str
    chave: str = ""

    def __post_init__(self):
        self.chave = _chave(self.nome_profissional, self.nome_proprietario, self.cidade)


def _normalizar(texto: str) -> str:
    if not texto:
        return ""
    txt = unicodedata.normalize("NFD", str(texto))
    txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
    return txt.strip().upper()


def _chave(profissional: str, proprietario: str, cidade: str) -> str:
    return f"{_normalizar(profissional)}|{_normalizar(proprietario)}|{_normalizar(cidade)}"


# ---------------------------------------------------------------------------
# Carregamento com detecção dinâmica de cabeçalhos
# ---------------------------------------------------------------------------

def carregar_excel(conteudo: bytes) -> tuple[openpyxl.Workbook, list[ObraRow], int]:
    """
    Carrega qualquer planilha Excel e extrai as obras usando detecção inteligente
    de cabeçalhos via Mistral AI (com fallback por palavras-chave).

    Retorna:
        (workbook, lista_de_obras, header_row_1indexed)
        onde header_row_1indexed é a linha do cabeçalho no workbook (base-1).
    """
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    # Lê as primeiras 10 linhas para análise
    max_row_sample = min(10, ws.max_row or 1)
    primeiras_linhas = [
        [cell for cell in row]
        for row in ws.iter_rows(min_row=1, max_row=max_row_sample, values_only=True)
    ]

    # Detecta estrutura via IA
    estrutura = detectar_estrutura_planilha(primeiras_linhas)

    header_row_0idx: int = int(estrutura.get("header_row_index") or 0)
    header_row_1idx: int = header_row_0idx + 1       # base-1 para openpyxl
    data_start_row: int = header_row_1idx + 1

    col_prof: int | None = estrutura.get("profissional")
    col_prop: int | None = estrutura.get("proprietario")
    col_cidade: int | None = estrutura.get("cidade")
    col_uf: int | None = estrutura.get("uf")
    col_end: int | None = estrutura.get("endereco")

    # Valida: precisa de ao menos profissional ou proprietário
    if col_prof is None and col_prop is None:
        cabecalhos_detectados = [str(c or "") for c in primeiras_linhas[header_row_0idx]] if primeiras_linhas else []
        raise ValueError(
            "Não foi possível identificar as colunas de profissional ou proprietário na planilha.\n"
            f"Cabeçalhos detectados na linha {header_row_1idx}: {cabecalhos_detectados}\n"
            "Verifique se o arquivo contém colunas com esses dados."
        )

    # Valores que indicam NULL em exportações de banco de dados (MySQL, PostgreSQL, etc.)
    _NULL_INDICATORS = {r"\n", r"\N", "null", "NULL", "none", "None", "NA", "N/A", "#N/A"}

    def _get_col(row: tuple, idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        val = str(row[idx] or "").strip()
        return "" if val in _NULL_INDICATORS else val

    obras: list[ObraRow] = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        profissional = _get_col(row, col_prof)
        proprietario = _get_col(row, col_prop)
        cidade = _get_col(row, col_cidade)
        endereco = _get_col(row, col_end)

        # UF: coluna dedicada ou extrai do endereço
        uf = _get_col(row, col_uf)
        if not uf and endereco:
            uf = extrair_uf_de_texto(endereco)

        if not profissional and not proprietario:
            logger.warning("Linha %d ignorada: profissional e proprietario vazios", row_idx)
            continue

        obras.append(
            ObraRow(
                row_index=row_idx,
                nome_profissional=profissional,
                nome_proprietario=proprietario,
                cidade=cidade,
                uf=uf,
                endereco=endereco,
            )
        )

    logger.info(
        "Excel carregado: %d obras | header_row=%d | mapeamento=%s",
        len(obras),
        header_row_1idx,
        {
            "profissional": col_prof,
            "proprietario": col_prop,
            "cidade": col_cidade,
            "uf": col_uf,
            "endereco": col_end,
        },
    )
    return wb, obras, header_row_1idx


# ---------------------------------------------------------------------------
# Enriquecimento
# ---------------------------------------------------------------------------

def enriquecer_excel(
    wb: openpyxl.Workbook,
    obras: list[ObraRow],
    contatos: list,
    header_row: int = 1,
) -> bytes:
    """
    Gera uma planilha limpa para download.
    A primeira linha da saída sempre contém o cabeçalho, seguida pelas linhas de dados.

    Args:
        wb: workbook original (para copiar colunas e valores)
        obras: lista de ObraRow extraídas
        contatos: resultados do scraper
        header_row: linha base-1 do cabeçalho no workbook original
    """
    ws_origem = wb.active
    mapa: dict[str, object] = {c.chave: c for c in contatos}
    mapa_por_row: dict[int, object] = {c.row_index: c for c in contatos}

    # Determina última coluna com conteúdo no cabeçalho
    original_max_col = ws_origem.max_column
    while original_max_col > 1 and not ws_origem.cell(header_row, original_max_col).value:
        original_max_col -= 1

    headers = [
        ws_origem.cell(header_row, col).value or f"Coluna {col}"
        for col in range(1, original_max_col + 1)
    ]

    wb_saida = openpyxl.Workbook()
    ws = wb_saida.active
    ws.title = "Leads Enriquecidos"
    ws.append(headers + COLUNAS_NOVAS)

    for saida_row, obra in enumerate(obras, start=2):
        contato = mapa_por_row.get(obra.row_index) or mapa.get(obra.chave)
        valores_originais = [
            ws_origem.cell(obra.row_index, col).value
            for col in range(1, original_max_col + 1)
        ]

        if contato is None:
            novos_valores = ["", "", "", "", "", "", "Nao processado"]
        elif contato.erro:
            novos_valores = ["", "", "", "", "", "", f"Erro: {contato.erro[:60]}"]
        else:
            tem_contato = bool(contato.tel_arq_1 or contato.tel_prop_1)
            status = "OK" if tem_contato else "Sem telefone cadastrado"
            novos_valores = [
                contato.tel_arq_1,
                contato.tel_arq_2,
                contato.email_arq,
                contato.tel_prop_1,
                contato.tel_prop_2,
                contato.email_prop,
                status,
            ]

        ws.append(valores_originais + novos_valores)
        ws.row_dimensions[saida_row].height = 34

    _formatar_planilha(ws, original_max_col, len(headers) + len(COLUNAS_NOVAS), len(obras) + 1)

    output = io.BytesIO()
    wb_saida.save(output)
    output.seek(0)
    logger.info("Excel enriquecido gerado.")
    return output.read()


# ---------------------------------------------------------------------------
# Formatação
# ---------------------------------------------------------------------------

def _formatar_planilha(ws, primeira_col_enriquecida: int, total_cols: int, total_rows: int) -> None:
    """Aplica formatação legível e consistente."""
    header_fill = PatternFill("solid", fgColor=COR_HEADER)
    new_header_fill = PatternFill("solid", fgColor=COR_HEADER_NOVO)
    zebra_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    border = Border(
        left=Side(style="thin", color=COR_BORDA),
        right=Side(style="thin", color=COR_BORDA),
        top=Side(style="thin", color=COR_BORDA),
        bottom=Side(style="thin", color=COR_BORDA),
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{max(total_rows, 1)}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32

    for col in range(1, total_cols + 1):
        cell = ws.cell(1, col)
        is_new = col > primeira_col_enriquecida
        cell.fill = new_header_fill if is_new else header_fill
        cell.font = Font(bold=True, color=COR_HEADER_NOVO_FONT if is_new else COR_HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in range(2, total_rows + 1):
        for col in range(1, total_cols + 1):
            cell = ws.cell(row, col)
            if row % 2 == 0:
                cell.fill = zebra_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if col > primeira_col_enriquecida:
                cell.font = Font(color="1F1F1F")

    for col in range(1, total_cols + 1):
        letter = get_column_letter(col)
        header = str(ws.cell(1, col).value or "")
        values = [str(ws.cell(row, col).value or "") for row in range(1, min(total_rows, 30) + 1)]
        longest = max([len(header), *(len(v) for v in values)], default=12)

        if col > primeira_col_enriquecida:
            width = min(max(longest + 3, 18), 32)
        elif "email" in header.lower() or "informa" in header.lower() or "feedback" in header.lower():
            width = min(max(longest + 2, 22), 38)
        elif "telefone" in header.lower() or "tel" in header.lower():
            width = 19
        else:
            width = min(max(longest + 2, 12), 28)

        ws.column_dimensions[letter].width = width


# ---------------------------------------------------------------------------
# Exportação formato Meetime
# ---------------------------------------------------------------------------

def normalizar_tel(tel: str) -> str:
    """Remove tudo que não for dígito. Ex: '(11) 9 9999-9999' -> '11999999999'."""
    if not tel:
        return ""
    return re.sub(r"\D", "", str(tel))


COLUNAS_MEETIME = [
    "Nome",
    "Tipo",
    "Telefone 1",
    "Telefone 2",
    "Email",
    "Empresa / Condominio",
    "Cidade",
    "UF",
    "Status",
]


def gerar_meetime_excel(obras: list[ObraRow], contatos: list) -> bytes:
    """
    Gera planilha no formato Meetime: 1 linha por contato, telefones normalizados.
    Aba 'Todos' com todos os contatos + uma aba por cidade.
    """
    mapa_por_row: dict[int, object] = {c.row_index: c for c in contatos}
    mapa_por_chave: dict[str, object] = {c.chave: c for c in contatos}

    def _linha(nome, tipo, tel1, tel2, email, empresa, cidade, uf, status, cidade_key):
        return {
            "Nome": nome,
            "Tipo": tipo,
            "Telefone 1": normalizar_tel(tel1),
            "Telefone 2": normalizar_tel(tel2),
            "Email": email or "",
            "Empresa / Condominio": empresa,
            "Cidade": cidade,
            "UF": uf,
            "Status": status,
            "_cidade_key": cidade_key,
        }

    def _montar_linhas_obra(obra: ObraRow) -> list[dict]:
        contato = mapa_por_row.get(obra.row_index) or mapa_por_chave.get(obra.chave)
        empresa = obra.endereco or ""
        cidade = obra.cidade
        uf = obra.uf
        cidade_key = _normalizar(cidade)

        if contato is None:
            resultado = []
            if obra.nome_profissional:
                resultado.append(_linha(obra.nome_profissional, "Arquiteto", "", "", "", empresa, cidade, uf, "Nao processado", cidade_key))
            if obra.nome_proprietario:
                resultado.append(_linha(obra.nome_proprietario, "Proprietario", "", "", "", empresa, cidade, uf, "Nao processado", cidade_key))
            return resultado

        if contato.erro:
            resultado = []
            if obra.nome_profissional:
                resultado.append(_linha(obra.nome_profissional, "Arquiteto", "", "", "", empresa, cidade, uf, "Erro", cidade_key))
            if obra.nome_proprietario:
                resultado.append(_linha(obra.nome_proprietario, "Proprietario", "", "", "", empresa, cidade, uf, "Erro", cidade_key))
            return resultado

        resultado = []
        nome_arq = contato.nome_arquiteto or obra.nome_profissional
        if nome_arq:
            status_arq = "Com telefone" if contato.tel_arq_1 else "Sem telefone"
            resultado.append(_linha(nome_arq, "Arquiteto", contato.tel_arq_1, contato.tel_arq_2, contato.email_arq, empresa, cidade, uf, status_arq, cidade_key))

        nome_prop = contato.nome_proprietario or obra.nome_proprietario
        if nome_prop:
            status_prop = "Com telefone" if contato.tel_prop_1 else "Sem telefone"
            resultado.append(_linha(nome_prop, "Proprietario", contato.tel_prop_1, contato.tel_prop_2, contato.email_prop, empresa, cidade, uf, status_prop, cidade_key))

        return resultado

    todas_linhas: list[dict] = []
    for obra in obras:
        todas_linhas.extend(_montar_linhas_obra(obra))

    wb = openpyxl.Workbook()
    ws_todos = wb.active
    ws_todos.title = "Todos"
    _escrever_aba_meetime(ws_todos, todas_linhas)

    cidades_vistas: list[tuple[str, str]] = []
    vistas: set[str] = set()
    for linha in todas_linhas:
        chave = linha["_cidade_key"]
        if chave and chave not in vistas:
            vistas.add(chave)
            cidades_vistas.append((chave, linha["Cidade"]))

    for chave, nome_cidade in cidades_vistas:
        linhas_cidade = [l for l in todas_linhas if l["_cidade_key"] == chave]
        if not linhas_cidade:
            continue
        titulo = (nome_cidade or chave or "Sem Cidade")[:31]
        ws_cidade = wb.create_sheet(title=titulo)
        _escrever_aba_meetime(ws_cidade, linhas_cidade)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("Excel Meetime gerado: %s contatos, %s cidades.", len(todas_linhas), len(cidades_vistas))
    return output.read()


def _escrever_aba_meetime(ws, linhas: list[dict]) -> None:
    ws.append(COLUNAS_MEETIME)
    for linha in linhas:
        ws.append([linha.get(c, "") for c in COLUNAS_MEETIME])
    _formatar_planilha_meetime(ws, len(linhas) + 1)


def _formatar_planilha_meetime(ws, total_rows: int) -> None:
    total_cols = len(COLUNAS_MEETIME)
    header_fill = PatternFill("solid", fgColor=COR_HEADER)
    zebra_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    border = Border(
        left=Side(style="thin", color=COR_BORDA),
        right=Side(style="thin", color=COR_BORDA),
        top=Side(style="thin", color=COR_BORDA),
        bottom=Side(style="thin", color=COR_BORDA),
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{max(total_rows, 1)}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32

    for col in range(1, total_cols + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=COR_HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    col_widths = {
        "Nome": 30, "Tipo": 14, "Telefone 1": 18, "Telefone 2": 18,
        "Email": 32, "Empresa / Condominio": 28, "Cidade": 18, "UF": 6, "Status": 16,
    }

    for row in range(2, total_rows + 1):
        ws.row_dimensions[row].height = 26
        for col in range(1, total_cols + 1):
            cell = ws.cell(row, col)
            if row % 2 == 0:
                cell.fill = zebra_fill
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = border

    for col, header in enumerate(COLUNAS_MEETIME, start=1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(header, 18)
