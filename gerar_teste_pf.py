"""
Gera arquivo de teste com arquitetos/profissionais individuais (PF)
para verificar se o Ver Mais retorna telefones para esse tipo de contato.

Formato: mesmo do export Mais Obras (Meus_favoritos.xls)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Meus favoritos"

# Cabeçalho no estilo Mais Obras
headers = [
    "Profissional",
    "Proprietario",
    "Cidade",
    "UF",
    "Endereco",
]

ws.append(headers)

# Arquitetos individuais (PF) de cidades do interior de SP
# Nomes genéricos mas plausíveis — formato: ARQ. NOME SOBRENOME
dados = [
    ("ARQ. CARLOS ROBERTO SILVA",      "MARCOS ANTONIO FERREIRA",    "SAO CARLOS",      "SP", "RUA DAS FLORES, 123"),
    ("ARQ. ANA PAULA MENDES",          "JOAO CARLOS LIMA",           "SAO CARLOS",      "SP", "AV. TRABALHADORES, 456"),
    ("ARQ. RENATO AUGUSTO COSTA",      "PATRICIA SOUZA",             "RIBEIRAO PRETO",  "SP", "RUA ARARAQUARA, 789"),
    ("ARQ. FERNANDA CRISTINA ALVES",   "ROBERTO CARLOS PEREIRA",     "RIBEIRAO PRETO",  "SP", "AV. BRASIL, 1234"),
    ("ARQ. PAULO HENRIQUE SANTOS",     "CLAUDIA REGINA OLIVEIRA",    "ARARAQUARA",      "SP", "RUA PROGRESSO, 321"),
    ("ARQ. MARCIA APARECIDA ROCHA",    "LUIZ ANTONIO MARTINS",       "ARARAQUARA",      "SP", "AV. SANTOS DUMONT, 567"),
    ("ARQ. THIAGO RODRIGUES",          "ANA MARIA CARVALHO",         "BAURU",           "SP", "RUA PARANA, 890"),
    ("ARQ. JULIANA BETTIM CASTRO",     "EDSON APARECIDO RIBEIRO",    "BAURU",           "SP", "AV. NAÇOES UNIDAS, 2345"),
    ("ARQ. RAFAEL MOREIRA DIAS",       "SILVIA HELENA CARDOSO",      "CAMPINAS",        "SP", "RUA JOSE PAULINO, 111"),
    ("ARQ. TATIANA LIMA NUNES",        "GUSTAVO HENRIQUE ARAUJO",    "CAMPINAS",        "SP", "AV. BARÃO DE ITAPURA, 222"),
    ("STUDIO ARQ. M. FERREIRA",        "MIGUEL ANTONIO FERREIRA",    "PIRACICABA",      "SP", "RUA QUINZE DE NOVEMBRO, 333"),
    ("ARQ. BEATRIZ GONCALVES",         "FLAVIO AUGUSTO MENDES",      "PIRACICABA",      "SP", "AV. INDEPENDENCIA, 444"),
]

for row in dados:
    ws.append(list(row))

# Formata cabeçalho
header_fill = PatternFill("solid", fgColor="56181B")
for col in range(1, len(headers) + 1):
    cell = ws.cell(1, col)
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFF7E2")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Ajusta largura
col_widths = [35, 35, 20, 5, 35]
from openpyxl.utils import get_column_letter
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

output_path = "ARQUITETOS_PF_TESTE.xlsx"
wb.save(output_path)
print(f"OK Arquivo gerado: {output_path}")
print(f"   {len(dados)} linhas - arquitetos PF de SP interior")
