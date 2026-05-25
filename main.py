"""
main.py — Mais Obras Enricher v1.2
Sem Playwright. Tudo via httpx direto na API interna do Mais Obras.
"""

import logging
import os
import subprocess
import tempfile
import uuid
from contextlib import asynccontextmanager

from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Security, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, Response
from fastapi.security.api_key import APIKeyHeader
from fastapi.staticfiles import StaticFiles

load_dotenv()

from excel_handler import carregar_excel, enriquecer_excel, gerar_meetime_excel
from scraper import MaisObrasScraper

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
)
logger = logging.getLogger(__name__)

MAISOBRAS_EMAIL    = os.getenv("MAISOBRAS_EMAIL", "")
MAISOBRAS_PASSWORD = os.getenv("MAISOBRAS_PASSWORD", "")
API_TOKEN          = os.getenv("API_TOKEN", "")
MAX_OBRAS          = int(os.getenv("MAX_OBRAS_PER_REQUEST", "1500"))
MISTRAL_API_KEY    = os.getenv("MISTRAL_API_KEY", "")

api_key_header = APIKeyHeader(name="X-API-Token", auto_error=False)

def verificar_token(token: str = Security(api_key_header)):
    if API_TOKEN and token != API_TOKEN:
        raise HTTPException(status_code=401, detail="Token de API inválido.")
    return token

scraper = MaisObrasScraper()
jobs: dict[str, dict] = {}

@asynccontextmanager
async def lifespan(app: FastAPI):
    await scraper.start()
    if MAISOBRAS_EMAIL and MAISOBRAS_PASSWORD:
        ok = await scraper.login(MAISOBRAS_EMAIL, MAISOBRAS_PASSWORD)
        if not ok:
            logger.error(
                "Autenticação inicial falhou. "
                "Use POST /reautenticar após verificar as credenciais."
            )
    else:
        logger.warning("MAISOBRAS_EMAIL/PASSWORD não configurados.")
    yield
    await scraper.stop()

app = FastAPI(
    title="Mais Obras Enricher",
    description=(
        "Enriquece o Excel de Favoritos do Mais Obras com telefones e e-mails "
        "via chamada direta à API interna da plataforma."
    ),
    version="1.2.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

app.mount("/assets", StaticFiles(directory="assets"), name="assets")
app.mount("/frontend", StaticFiles(directory="frontend-source"), name="frontend")


@app.get("/", response_class=HTMLResponse, tags=["Interface"])
async def index():
    with open("frontend-source/Auto ECO.html", "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())


@app.get("/health", tags=["Sistema"])
async def health():
    return {
        "status": "ok",
        "scraper_autenticado": scraper._authenticated,
        "max_obras_por_request": MAX_OBRAS,
        "versao": "1.2.0",
        "modo": "httpx (sem browser)",
    }


def _converter_xls_para_xlsx(conteudo_xls: bytes) -> bytes:
    """Converte .xls legado para .xlsx via LibreOffice."""
    import io
    import shutil

    from openpyxl import Workbook
    from openpyxl.styles import Font

    def converter_com_xlrd() -> bytes:
        try:
            from python_calamine import load_workbook as load_calamine_workbook

            book = load_calamine_workbook(io.BytesIO(conteudo_xls))
            sheet = book.get_sheet_by_index(0)
            rows = list(sheet.iter_rows())
            wb = Workbook()
            ws = wb.active
            ws.title = (book.sheet_names[0] if book.sheet_names else "Planilha")[:31]

            for row_idx, row in enumerate(rows, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            if ws.max_row >= 4:
                for cell in ws[4]:
                    cell.font = Font(bold=True)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
        except Exception as calamine_error:
            logger.warning(f"Falha ao converter .xls com python-calamine: {calamine_error}")

        import xlrd

        book = xlrd.open_workbook(file_contents=conteudo_xls, formatting_info=False)
        sheet = book.sheet_by_index(0)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet.name[:31] or "Planilha"

        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                    value = int(value)
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

        if ws.max_row >= 4:
            for cell in ws[4]:
                cell.font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    if not shutil.which("libreoffice"):
        return converter_com_xlrd()

    with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp_in:
        tmp_in.write(conteudo_xls)
        tmp_path = tmp_in.name

    out_dir = tempfile.mkdtemp()
    try:
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx",
             "--outdir", out_dir, tmp_path],
            capture_output=True, timeout=60, check=True,
        )
        import os, glob
        xlsx_files = glob.glob(os.path.join(out_dir, "*.xlsx"))
        if not xlsx_files:
            raise RuntimeError("LibreOffice não gerou arquivo .xlsx")
        with open(xlsx_files[0], "rb") as f:
            return f.read()
    finally:
        import os, shutil
        try:
            os.unlink(tmp_path)
            shutil.rmtree(out_dir, ignore_errors=True)
        except Exception:
            pass


def _converter_csv_para_xlsx(conteudo_csv: bytes) -> bytes:
    """Converte CSV exportado do Mais Obras para .xlsx, detectando encoding automaticamente."""
    import csv

    texto: str | None = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            texto = conteudo_csv.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise ValueError("Nao foi possivel decodificar o CSV.")

    leitor = csv.reader(io.StringIO(texto))
    linhas = list(leitor)
    if not linhas:
        raise ValueError("CSV vazio.")

    wb_csv = openpyxl.Workbook()
    ws_csv = wb_csv.active
    ws_csv.title = "Planilha"

    # Insere os dados diretamente — a detecção de cabeçalho é feita pela IA
    for linha in linhas:
        ws_csv.append(linha)

    output = io.BytesIO()
    wb_csv.save(output)
    output.seek(0)
    return output.read()


async def _processar_job(job_id: str, filename: str, conteudo: bytes, modo_meetime: bool = False):
    job = jobs[job_id]
    try:
        job.update(status="running", message="Preparando arquivo...", current_line="Preparando arquivo...")

        if modo_meetime:
            job.update(current_line="Modo Meetime ativado: exportara 1 linha por contato, separado por cidade.")

        ext = (filename or "").lower().rsplit(".", 1)[-1]
        if ext == "xls":
            job.update(message="Convertendo .xls para .xlsx...", current_line="Convertendo .xls para .xlsx...")
            conteudo = _converter_xls_para_xlsx(conteudo)
        elif ext == "csv":
            job.update(message="Convertendo .csv para .xlsx...", current_line="Convertendo .csv para .xlsx...")
            conteudo = _converter_csv_para_xlsx(conteudo)

        job.update(message="Lendo planilha e detectando colunas com IA...", current_line="Detectando colunas via IA...")
        wb, obras, header_row = carregar_excel(conteudo)

        if not obras:
            raise ValueError("Nenhuma obra encontrada no arquivo.")
        if len(obras) > MAX_OBRAS:
            raise ValueError(f"O arquivo tem {len(obras)} obras, limite é {MAX_OBRAS}.")

        job.update(
            total=len(obras),
            processed=0,
            message=f"{len(obras)} obras encontradas.",
            current_line=f"{len(obras)} obras encontradas. Iniciando consultas...",
        )

        def progress_callback(obra, contato):
            processed = int(job.get("processed", 0)) + 1
            nome = obra.nome_profissional or obra.nome_proprietario or "sem nome"
            tem_tel = bool(contato.tel_arq_1 or contato.tel_prop_1)
            status = "OK" if tem_tel else "sem telefone"
            summary = f"Linha {obra.row_index} | {processed}/{len(obras)} | {nome[:50]} | {status}"

            # Cabeçalho da linha
            job["logs"].append({"text": f"─── {summary}", "kind": "ok" if tem_tel else "warn"})
            # Logs detalhados do scraper (ARQ/PROP/perfil/ver_mais/IA)
            for msg in getattr(contato, "log_messages", []):
                kind = "ok" if "✓" in msg else ("warn" if "sem telefone" in msg or "empresa" in msg.lower() else "err" if "erro" in msg.lower() else "")
                job["logs"].append({"text": msg, "kind": kind})

            job.update(
                processed=processed,
                current_row=obra.row_index,
                current_contact=nome,
                current_line=summary,
            )

        contatos = await scraper.processar_lote(obras, progress_callback=progress_callback)

        job.update(message="Gerando Excel final...", current_line="Gerando Excel final...")
        if modo_meetime:
            excel_bytes = gerar_meetime_excel(obras, contatos)
            nome_saida = (filename or "favoritos").rsplit(".", 1)[0] + "_meetime.xlsx"
        else:
            excel_bytes = enriquecer_excel(wb, obras, contatos, header_row=header_row)
            nome_saida = (filename or "favoritos").rsplit(".", 1)[0] + "_enriquecido.xlsx"

        sucesso = sum(1 for c in contatos if c.tel_arq_1 or c.tel_prop_1)
        job.update(
            status="done",
            result=excel_bytes,
            output_name=nome_saida,
            current_line=f"Concluido: {sucesso}/{len(contatos)} linhas com telefone encontrado.",
            message="Concluido.",
        )
    except Exception as e:
        logger.exception("Falha no job de enriquecimento")
        job.update(status="failed", error=str(e), current_line=f"Erro: {e}", message="Falha.")


@app.post("/enriquecer_async", tags=["Enriquecimento"])
async def enriquecer_async(
    background_tasks: BackgroundTasks,
    arquivo: UploadFile = File(...),
    modo_meetime: str = Form("0"),
    _token: str = Security(verificar_token),
):
    if not scraper._authenticated:
        raise HTTPException(503, "Scraper não autenticado.")

    ext = (arquivo.filename or "").lower().rsplit(".", 1)[-1]
    if ext not in ("xls", "xlsx", "xlsm", "csv"):
        raise HTTPException(400, "Envie um arquivo .xls, .xlsx ou .csv exportado do Mais Obras.")

    conteudo = await arquivo.read()
    if not conteudo:
        raise HTTPException(400, "Arquivo vazio.")

    job_id = uuid.uuid4().hex
    jobs[job_id] = {
        "status": "queued",
        "filename": arquivo.filename or "favoritos",
        "total": 0,
        "processed": 0,
        "current_row": None,
        "current_contact": "",
        "current_line": "Na fila para processamento...",
        "message": "Na fila.",
        "result": None,
        "output_name": "",
        "error": "",
        "logs": [],   # lista de {"text": str, "kind": str} — drenada pelo frontend
    }
    meetime = modo_meetime.strip() in ("1", "true", "on", "yes")
    background_tasks.add_task(_processar_job, job_id, arquivo.filename or "favoritos", conteudo, meetime)
    return {"job_id": job_id}


@app.get("/progresso/{job_id}", tags=["Enriquecimento"])
async def progresso(job_id: str):
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job não encontrado.")
    return {
        "job_id": job_id,
        "status": job.get("status"),
        "total": job.get("total", 0),
        "processed": job.get("processed", 0),
        "current_row": job.get("current_row"),
        "current_contact": job.get("current_contact"),
        "current_line": job.get("current_line"),
        "message": job.get("message"),
        "error": job.get("error"),
        "logs": job.get("logs", []),
    }


@app.get("/resultado/{job_id}", tags=["Enriquecimento"])
async def resultado(job_id: str):
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job não encontrado.")
    if job.get("status") != "done" or not job.get("result"):
        raise HTTPException(409, "Resultado ainda não está pronto.")

    return Response(
        content=job["result"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{job.get("output_name") or "Meus_favoritos_enriquecido.xlsx"}"'},
    )


@app.post(
    "/enriquecer",
    tags=["Enriquecimento"],
    response_class=Response,
    summary="Enriquece o Excel com telefones e e-mails",
    responses={
        200: {"description": "Excel enriquecido (.xlsx)"},
        400: {"description": "Arquivo inválido"},
        401: {"description": "Token inválido"},
        429: {"description": "Limite de obras excedido"},
        503: {"description": "Scraper não autenticado"},
    },
)
async def enriquecer(
    arquivo: UploadFile = File(
        ...,
        description="Arquivo 'Meus_favoritos.xls' exportado do Mais Obras"
    ),
    modo_meetime: str = Form("0"),
    _token: str = Security(verificar_token),
):
    """
    Recebe o Excel de favoritos exportado do Mais Obras,
    consulta o endpoint /pesquisa_perfil para cada obra,
    e devolve o Excel com 7 novas colunas de contato.
    """
    if not scraper._authenticated:
        raise HTTPException(
            503,
            "Scraper não autenticado. "
            "Configure MAISOBRAS_EMAIL e MAISOBRAS_PASSWORD e chame POST /reautenticar.",
        )

    ext = (arquivo.filename or "").lower().rsplit(".", 1)[-1]
    if ext not in ("xls", "xlsx", "xlsm", "csv"):
        raise HTTPException(400, "Envie um arquivo .xls, .xlsx ou .csv exportado do Mais Obras.")

    conteudo = await arquivo.read()
    if not conteudo:
        raise HTTPException(400, "Arquivo vazio.")

    # Converte formatos legados/alternativos para .xlsx
    if ext == "xls":
        try:
            conteudo = _converter_xls_para_xlsx(conteudo)
        except Exception as e:
            raise HTTPException(400, f"Erro ao converter .xls para .xlsx: {e}")
    elif ext == "csv":
        try:
            conteudo = _converter_csv_para_xlsx(conteudo)
        except Exception as e:
            raise HTTPException(400, f"Erro ao converter .csv para .xlsx: {e}")

    # Carrega e detecta estrutura do Excel via IA
    try:
        wb, obras, header_row = carregar_excel(conteudo)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"Erro ao ler arquivo: {e}")

    if not obras:
        raise HTTPException(400, "Nenhuma obra encontrada no arquivo.")

    if len(obras) > MAX_OBRAS:
        raise HTTPException(
            429,
            f"O arquivo tem {len(obras)} obras, limite é {MAX_OBRAS}. "
            "Divida em arquivos menores ou aumente MAX_OBRAS_PER_REQUEST no Railway.",
        )

    logger.info(f"Enriquecendo {len(obras)} obras de '{arquivo.filename}'...")

    # Coleta contatos via API
    contatos = await scraper.processar_lote(obras)

    # Gera Excel enriquecido
    meetime = modo_meetime.strip() in ("1", "true", "on", "yes")
    try:
        if meetime:
            excel_bytes = gerar_meetime_excel(obras, contatos)
        else:
            excel_bytes = enriquecer_excel(wb, obras, contatos, header_row=header_row)
    except Exception as e:
        raise HTTPException(500, f"Erro ao gerar arquivo: {e}")

    sucesso = sum(1 for c in contatos if c.tel_arq_1 or c.tel_prop_1)
    logger.info(f"Concluído: {sucesso}/{len(contatos)} obras com telefone encontrado.")

    sufixo = "_meetime.xlsx" if meetime else "_enriquecido.xlsx"
    nome_saida = (arquivo.filename or "favoritos").rsplit(".", 1)[0] + sufixo

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_saida}"'},
    )


@app.post("/reautenticar", tags=["Sistema"])
async def reautenticar(_token: str = Security(verificar_token)):
    """Força novo login no Mais Obras (usar quando a sessão expirar)."""
    if not MAISOBRAS_EMAIL or not MAISOBRAS_PASSWORD:
        raise HTTPException(400, "Configure MAISOBRAS_EMAIL e MAISOBRAS_PASSWORD.")
    ok = await scraper.login(MAISOBRAS_EMAIL, MAISOBRAS_PASSWORD)
    if ok:
        return {"status": "autenticado"}
    raise HTTPException(503, "Falha na autenticação. Verifique as credenciais.")


@app.get("/debug/ver_mais_js", tags=["Debug"])
async def debug_ver_mais_js(obra_id: str = ""):
    """
    Usa Playwright para inspecionar o JS do Ver Mais diretamente na SPA do Mais Obras.

    Como a URL nunca muda (SPA), navega para /pesquisa_obras, carrega os scripts
    JS da página e procura o handler pesquisa_contatos_api. Opcionalmente recebe
    um obra_id para tentar acionar o carregamento dessa obra no painel.

    Retorna trechos de JS relevantes + requests capturadas durante a interação.
    """
    if not scraper._pw_page:
        return {"erro": "Playwright não disponível — reinicie o servidor."}

    base_url = os.getenv("MAISOBRAS_BASE_URL", "https://www.maisobras.online")

    async with scraper._pw_lock:
        page = scraper._pw_page
        captured_requests: list = []

        async def on_request(request):
            # Captura QUALQUER request de API (não só pesquisa_contatos)
            url = request.url
            if any(k in url for k in ("/api/", "pesquisa", "contatos", "ver_mais", "perfil")):
                try:
                    body = request.post_data or ""
                except Exception:
                    body = ""
                captured_requests.append({
                    "url": url,
                    "method": request.method,
                    "post_data": body[:500],
                })

        page.on("request", on_request)

        try:
            # Navega para a pesquisa de obras (SPA — URL fixa)
            logger.info("[debug/ver_mais_js] Carregando pesquisa_obras...")
            await page.goto(base_url + "/pesquisa_obras", wait_until="networkidle", timeout=25000)
            await page.wait_for_timeout(3000)

            page_url_atual = page.url
            page_title = await page.title()

            # Se um obra_id foi passado, tenta acionar o carregamento via JS da própria SPA
            if obra_id.strip():
                logger.info("[debug/ver_mais_js] Tentando carregar obra_id=%s via JS da SPA", obra_id)
                load_result = await page.evaluate("""async (obraId) => {
                    // Tenta as funções JS comuns de SPAs CodeIgniter
                    const fns = ['carregarObra', 'loadObra', 'abrirObra', 'verObra',
                                  'detalheObra', 'mostrarObra', 'getObra', 'openObra'];
                    for (const fn of fns) {
                        if (typeof window[fn] === 'function') {
                            try { window[fn](obraId); return {chamou: fn}; }
                            catch(e) {}
                        }
                    }
                    // Tenta acionar clicando em elemento com o id
                    const els = [...document.querySelectorAll('[data-id="' + obraId + '"], [data-obra="' + obraId + '"]')];
                    if (els.length > 0) { els[0].click(); return {clicou: els[0].tagName}; }
                    return {tentativas: fns, resultado: 'nenhuma funcao encontrada'};
                }""", obra_id.strip())
                logger.info("[debug/ver_mais_js] load_result: %s", load_result)
                await page.wait_for_timeout(2000)

            # Coleta TODOS os scripts externos carregados
            script_srcs = await page.evaluate("""() => {
                return [...document.querySelectorAll('script[src]')].map(s => s.src);
            }""")

            # Coleta scripts inline relevantes
            inline_relevantes = await page.evaluate("""() => {
                return [...document.querySelectorAll('script:not([src])')]
                    .map(s => s.textContent)
                    .filter(t =>
                        t.includes('pesquisa_contatos') || t.includes('verMais') ||
                        t.includes('ver_mais') || t.includes('SequentialId') ||
                        t.includes('sequence_id') || t.includes('pesquisa_perfil')
                    );
            }""")

            # Tenta encontrar botões/elementos "Ver Mais" já renderizados
            ver_mais_elements = await page.evaluate("""() => {
                return [...document.querySelectorAll('*')]
                    .filter(el => {
                        try {
                            const txt = el.textContent || '';
                            const onclick = el.getAttribute('onclick') || '';
                            const cls = String(el.className || '');
                            return (txt.trim().toLowerCase() === 'ver mais') ||
                                   (txt.trim().toLowerCase().includes('ver mais') && el.children.length === 0) ||
                                   onclick.includes('pesquisa_contatos') || onclick.includes('verMais') ||
                                   cls.includes('ver-mais') || cls.includes('ver_mais');
                        } catch(e) { return false; }
                    })
                    .slice(0, 15)
                    .map(el => ({
                        tag: el.tagName,
                        text: el.textContent.trim().substring(0, 100),
                        onclick: el.getAttribute('onclick') || '',
                        className: String(el.className || '').substring(0, 120),
                        dataAttrs: JSON.stringify(el.dataset || {}),
                    }));
            }""")

            # Busca o JS relevante nos arquivos externos
            js_sources_relevantes = []
            keywords = ("pesquisa_contatos_api", "verMais", "ver_mais", "SequentialId",
                        "sequence_id", "pesquisa_perfil")
            skip = ("jquery", "bootstrap", "font-awesome", "popper", "moment",
                    "datatables", "select2", "chart")

            for src in script_srcs:
                if not src:
                    continue
                if any(s in src.lower() for s in skip):
                    continue
                try:
                    js_content = await page.evaluate(
                        """async (url) => {
                            try { const r = await fetch(url); return await r.text(); }
                            catch(e) { return ''; }
                        }""", src,
                    )
                    for kw in keywords:
                        idx = js_content.find(kw)
                        if idx >= 0:
                            trecho = js_content[max(0, idx - 400): idx + 800]
                            js_sources_relevantes.append({
                                "src": src,
                                "keyword": kw,
                                "trecho": trecho,
                            })
                            # Continua buscando outras keywords no mesmo arquivo
                except Exception as e:
                    logger.debug("Erro ao buscar script %s: %s", src[:60], e)

            # Estado de login do Playwright
            is_logged_in = await page.evaluate("""() => {
                const body = document.body.innerHTML;
                return {
                    hasLogout: body.includes('logout') || body.includes('sair'),
                    hasPesquisaObras: body.includes('pesquisa_obras'),
                    hasMeusF: body.includes('Meus Favoritos') || body.includes('favoritos'),
                    bodyLen: body.length,
                    currentUrl: location.href,
                };
            }""")

            return {
                "page_url": page_url_atual,
                "page_title": page_title,
                "login_state": is_logged_in,
                "requests_capturadas": captured_requests,
                "ver_mais_elements": ver_mais_elements,
                "inline_scripts_relevantes": inline_relevantes,
                "js_externos_relevantes": js_sources_relevantes,
                "todos_script_srcs": script_srcs,
            }

        except Exception as e:
            logger.exception("Erro em debug/ver_mais_js")
            return {"erro": str(e), "requests_capturadas": captured_requests}
        finally:
            try:
                page.remove_listener("request", on_request)
            except Exception:
                pass
            try:
                await page.goto(base_url + "/pesquisa_obras", wait_until="domcontentloaded", timeout=10000)
            except Exception:
                pass


@app.get("/debug/ver_mais_test", tags=["Debug"])
async def debug_ver_mais_test(nome: str = "LIDIANA SOARES BREITSCHAFT", uf: str = "SP"):
    """
    Testa a chamada Ver Mais via httpx E via Playwright para o mesmo contato.
    Captura os headers reais enviados por cada método.
    Útil para diagnosticar por que a API retorna [] para todos os contatos.
    """
    import json as _json

    base_url = os.getenv("MAISOBRAS_BASE_URL", "https://www.maisobras.online")
    endpoint = base_url + "/api/pesquisa_contatos_api"
    payload = {"nome": nome, "cpfcnpj": "", "uf": uf, "ccp": "1"}
    contato_json = _json.dumps(payload, ensure_ascii=False)

    resultado = {
        "nome_testado": nome,
        "uf": uf,
        "payload_enviado": payload,
        "httpx": None,
        "playwright": None,
    }

    # ── 1. Testa via httpx (mesma sessão do pesquisa_perfil) ───────────
    try:
        r = await scraper._client.post(endpoint, data={"contato": contato_json})
        resultado["httpx"] = {
            "status": r.status_code,
            "body": r.text[:500],
            "cookies_enviados": {k: v[:30] + "..." for k, v in dict(scraper._client.cookies).items()},
            "headers_enviados": dict(r.request.headers),
        }
    except Exception as e:
        resultado["httpx"] = {"erro": str(e)}

    # ── 2. Testa via Playwright com interceptação dos headers reais ─────
    if scraper._pw_page:
        async with scraper._pw_lock:
            page = scraper._pw_page
            captured = {}

            async def interceptar(route, request):
                if "pesquisa_contatos_api" in request.url:
                    captured["request_url"] = request.url
                    captured["request_headers"] = dict(request.headers)
                    captured["request_post_data"] = request.post_data or ""
                    response = await route.fetch()
                    captured["response_status"] = response.status
                    captured["response_body"] = await response.text()
                    captured["response_headers"] = dict(response.headers)
                    await route.fulfill(response=response)
                else:
                    await route.continue_()

            await page.route("**/*", interceptar)
            try:
                pw_result = await page.evaluate(
                    """async ([url, contato]) => {
                        const jq = (typeof jQuery !== 'undefined') ? jQuery : (typeof $ !== 'undefined' ? $ : null);
                        if (jq) {
                            return new Promise((resolve) => {
                                jq.ajax({
                                    url: url, type: 'POST',
                                    data: {contato: contato},
                                    dataType: 'json',
                                    headers: {'X-Requested-With': 'XMLHttpRequest'},
                                })
                                .done(d => resolve({ok: true, body: JSON.stringify(d), method: 'jquery'}))
                                .fail(x => resolve({ok: false, status: x.status, body: x.responseText, method: 'jquery-err'}));
                            });
                        }
                        return {erro: 'jQuery nao disponivel'};
                    }""",
                    [endpoint, contato_json],
                )
                resultado["playwright"] = {
                    "js_result": pw_result,
                    "interceptado": captured,
                    "page_url": page.url,
                    "cookies": [{"name": c["name"], "value": c["value"][:20] + "..."} for c in await page.context.cookies()],
                }
            except Exception as e:
                resultado["playwright"] = {"erro": str(e), "interceptado": captured}
            finally:
                await page.unroute("**/*")
    else:
        resultado["playwright"] = {"erro": "Playwright nao disponivel"}

    return resultado


@app.get("/debug/perfil_test", tags=["Debug"])
async def debug_perfil_test(
    nome: str = "LIDIANA SOARES BREITSCHAFT",
    uf: str = "SP",
    tipo: str = "Proprietário",
    cidade: str = "",
):
    """
    Chama /pesquisa_perfil e retorna a resposta RAW completa.
    Mostra todos os perfis encontrados com cidade e telefones — útil para verificar
    se estamos pegando o perfil certo para nomes comuns.
    """
    import json as _json

    base_url = os.getenv("MAISOBRAS_BASE_URL", "https://www.maisobras.online")
    payload = {"contato": nome, "tipo": tipo, "cpf_cnpj": "", "uf": uf}

    try:
        r = await scraper._client.post(base_url + "/pesquisa_perfil", data=payload)
        raw = r.text
        try:
            data = r.json()
        except Exception:
            data = None

        perfil = (data or {}).get("perfil") or []
        resumo = [
            {
                "idx": i,
                "cidade": p.get("cidade", "?"),
                "uf": p.get("uf", "?"),
                "cbo": p.get("cbo", "?"),
                "telefones": p.get("telefones", ""),
                "cpfcnpj_presente": bool(p.get("cpfcnpj")),
            }
            for i, p in enumerate(perfil)
        ]

        # Qual escolheríamos com a lógica atual (filtro por cidade)
        escolhido_idx = 0
        if cidade and len(perfil) > 1:
            from scraper import _normalizar
            cidade_norm = _normalizar(cidade)
            for i, p in enumerate(perfil):
                if _normalizar(p.get("cidade") or "") == cidade_norm:
                    escolhido_idx = i
                    break

        return {
            "nome": nome, "tipo": tipo, "uf": uf, "cidade_filtro": cidade,
            "status_http": r.status_code,
            "total_perfis": len(perfil),
            "escolhido_idx": escolhido_idx,
            "perfis": resumo,
            "emails": [(data or {}).get("emails") or []],
            "raw_body_truncado": raw[:800],
        }
    except Exception as e:
        return {"erro": str(e)}
