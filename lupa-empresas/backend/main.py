"""Lupa de Empresas - FastAPI backend.

Serve a API (busca, dados da empresa, funcionarios do LinkedIn) e tambem os
arquivos estaticos do frontend, tudo em http://localhost:8010.
"""

import os

# Carrega .env (se existir) ANTES de importar modulos que leem env no import.
try:
    from dotenv import load_dotenv

    _ENV_PATH = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"
    )
    load_dotenv(_ENV_PATH)
except Exception:
    pass

import asyncio
import io
import re
import sys
import uuid

from fastapi import Body, FastAPI, File, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

import assertiva
import brasilapi
import casadosdados
import config_store
import custos
import donodozap
import meetime
import navlog
import sheet_reader
import linkedin_scraper
import mkbuscas
import serasa

# Base local de CPF (JBR_PF) — modulo compartilhado em ../../jbr_base.
_JBR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "jbr_base",
)
if _JBR not in sys.path:
    sys.path.insert(0, _JBR)
try:
    import cpf_lookup
except Exception:
    cpf_lookup = None

# Base local de CNPJ (Dados Abertos RFB) — modulo em ../../cnpj_base.
_CNPJ = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "cnpj_base",
)
if _CNPJ not in sys.path:
    sys.path.insert(0, _CNPJ)
try:
    import cnpj_lookup
except Exception:
    cnpj_lookup = None


def _cnpj_local() -> bool:
    return bool(cnpj_lookup and cnpj_lookup.ready())

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")

app = FastAPI(title="Lupa de Empresas", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── SERVIÇO DE DADOS (local) ──────────────────────────────────────────────
# A autenticação (login/usuários) vive no app-online (Render). Aqui só validamos
# o SEGREDO DE PROXY: só o app-online (que conhece o segredo) pode chamar as rotas
# de dados através do túnel. Sem PROXY_SECRET definido → modo dev aberto (localhost).
_PROXY_SECRET = os.environ.get("PROXY_SECRET", "").strip()


@app.get("/api/health")
async def health():
    return {"ok": True, "service": "capiblu-data", "protegido": bool(_PROXY_SECRET)}


@app.middleware("http")
async def _proxy_guard(request, call_next):
    path = request.url.path
    if request.method == "OPTIONS" or not path.startswith("/api/") or path == "/api/health":
        return await call_next(request)
    if _PROXY_SECRET and request.headers.get("x-proxy-secret", "") != _PROXY_SECRET:
        return JSONResponse({"detail": "Acesso negado (segredo de proxy)."}, status_code=401)
    return await call_next(request)


def _cpf_ready() -> bool:
    return bool(cpf_lookup and cpf_lookup.ready())


def _enrich_qsa_cpf(company: dict) -> None:
    """Enriquece cada socio do QSA com o CPF completo, quando resolvivel.

    Cruza nome + 6 digitos do meio da mascara (ex.: '***912137**') na base JBR.
    """
    if not _cpf_ready():
        return
    qsa = company.get("qsa") or []
    for socio in qsa:
        if not isinstance(socio, dict):
            continue
        nome = socio.get("nome_socio") or ""
        mask = socio.get("cnpj_cpf_do_socio") or ""
        # So PF: mascara de CPF tem 6 digitos (CNPJ de socio PJ tem outro formato).
        try:
            res = cpf_lookup.resolve_socio(nome, mask)
        except Exception:
            continue
        socio["cpf_status"] = res.get("status")
        if res.get("status") == "resolved":
            socio["cpf_completo"] = res.get("cpf")
            p = res.get("pessoa") or {}
            socio["nascimento"] = p.get("nascimento")
            socio["sexo"] = p.get("sexo")


@app.get("/api/person/name-search")
def person_name_search(
    q: str = "", broad: bool = False, limit: int = 40, offset: int = 0
):
    """Busca pessoas por nome na base JBR. broad=true usa LIKE (nomes compostos).

    Retorna 'total' = total de matches disponíveis (para 'Ver mais'/'Buscar todos').
    """
    if not _cpf_ready():
        return {"status": "unavailable", "message": "Base JBR ainda carregando."}
    if not q.strip():
        return {"status": "error", "message": "Parâmetro q obrigatório."}
    try:
        q = q.strip()
        if broad:
            pessoas = cpf_lookup.by_name_broad(q, limit=limit, offset=offset)
            total = cpf_lookup.count_name_broad(q)
        else:
            pessoas = cpf_lookup.by_name(q, limit=limit)
            total = len(pessoas)
        return {
            "status": "ok",
            "total": total,
            "returned": len(pessoas),
            "offset": offset,
            "broad": broad,
            "pessoas": pessoas,
        }
    except Exception as exc:
        return {"status": "error", "message": str(exc)[:120]}


@app.get("/api/person/{cpf}/mk")
async def person_mk(cpf: str):
    """Dados completos da Mk Buscas (intelgrax-cpfv2) para um CPF."""
    if not mkbuscas.enabled():
        return {"status": "unavailable", "message": "Mk não configurada."}
    return await mkbuscas.consulta_cpf(cpf)


@app.get("/api/cnpj/lookup")
def cnpj_lookup_list(tipo: str = "cnae"):
    """Lista de códigos de apoio (cnae|natureza|municipio) para os selects do front."""
    if not cnpj_lookup:
        return {"status": "unavailable", "itens": []}
    if tipo not in ("cnae", "natureza", "municipio", "pais", "qualificacao", "motivo"):
        return {"status": "error", "message": "tipo inválido", "itens": []}
    return {"status": "ok", "tipo": tipo, "itens": cnpj_lookup.list_lookup(tipo)}


@app.get("/api/phone/{phone}/reverse")
async def phone_reverse(phone: str):
    """Telefone reverso (WorkAPI intelgrax-tel): CPFs/CNPJs atrelados ao número."""
    return await mkbuscas.consulta_telefone(phone)


@app.get("/api/phone/{phone}/pertence/{doc}")
async def phone_pertence(phone: str, doc: str):
    """Valida se um CPF/CNPJ está atrelado a um telefone (validação de contato)."""
    return await mkbuscas.telefone_pertence(phone, doc)


@app.get("/api/phone/{phone}/donodozap")
async def phone_donodozap(phone: str, nome: str = ""):
    """Valida se um telefone pertence a determinada pessoa via DonoDoZap."""
    return await donodozap.consultar(phone, nome)


# ---- Busca Assertiva (API Localize V3) ----

@app.get("/api/assertiva/status")
async def assertiva_status():
    """Diz se a integração Assertiva está configurada (sem expor credenciais)."""
    return {"enabled": assertiva.enabled(), "finalidade_padrao": assertiva.DEFAULT_FINALIDADE}


@app.get("/api/assertiva/cpf")
async def assertiva_cpf(cpf: str, finalidade: int | None = None):
    return await assertiva.consulta_cpf(cpf, finalidade)


@app.get("/api/assertiva/cnpj")
async def assertiva_cnpj(cnpj: str, finalidade: int | None = None):
    return await assertiva.consulta_cnpj(cnpj, finalidade)


@app.get("/api/assertiva/telefone")
async def assertiva_telefone(telefone: str, finalidade: int | None = None):
    return await assertiva.consulta_telefone(telefone, finalidade)


@app.get("/api/assertiva/email")
async def assertiva_email(email: str, finalidade: int | None = None):
    return await assertiva.consulta_email(email, finalidade)


@app.post("/api/assertiva/nome")
async def assertiva_nome(payload: dict = Body(default={})):
    """Busca por nome/razão social e/ou endereço (Localize nome-endereco)."""
    filtros = payload.get("filtros") or payload or {}
    finalidade = payload.get("finalidade")
    return await assertiva.busca_nome_endereco(filtros, finalidade)


@app.get("/api/search")
async def search(q: str = ""):
    """Busca empresas por CNPJ (exato) ou nome. NAO faz scraping de LinkedIn."""
    return await brasilapi.search_companies(q)


@app.get("/api/person/{cpf}")
async def person_by_cpf(cpf: str):
    """Consulta identidade por CPF exato na base JBR."""
    if not _cpf_ready():
        return {"status": "unavailable", "message": "Base de CPF ainda carregando ou indisponivel."}
    p = cpf_lookup.by_cpf(cpf)
    return {"status": "ok" if p else "not_found", "pessoa": p}


@app.get("/api/person/resolve/")
async def person_resolve(name: str = "", mask: str = ""):
    """Resolve CPF por nome (+ mascara opcional). Usado para socios/funcionarios."""
    if not _cpf_ready():
        return {"status": "unavailable", "message": "Base de CPF ainda carregando ou indisponivel."}
    return cpf_lookup.resolve_socio(name, mask)


@app.get("/api/person/{cpf}/contacts")
async def person_contacts(cpf: str):
    """Telefones/emails de uma PF por CPF, via Serasa Infomais (on-demand)."""
    if not serasa.enabled():
        return {"status": "unavailable", "message": "Serasa nao configurada (defina SERASA_CLIENT_ID/SECRET)."}
    return await serasa.enrich_person(cpf)


@app.get("/api/company/{cnpj}/contacts")
async def company_contacts(cnpj: str):
    """Telefones/emails de uma PJ por CNPJ, via Serasa Infomais (on-demand)."""
    if not serasa.enabled():
        return {"status": "unavailable", "message": "Serasa nao configurada (defina SERASA_CLIENT_ID/SECRET)."}
    return await serasa.enrich_company(cnpj)


@app.get("/api/company/{cnpj}")
async def company(cnpj: str):
    """Dados completos da empresa via BrasilAPI. NAO faz scraping."""
    try:
        data = await brasilapi.fetch_company(cnpj)
        _enrich_qsa_cpf(data)
        return {"status": "ok", "company": data}
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"status": "error", "message": str(exc)})
    except LookupError as exc:
        return JSONResponse(status_code=404, content={"status": "error", "message": str(exc)})
    except Exception:
        return JSONResponse(
            status_code=502,
            content={"status": "error", "message": "Falha ao consultar a BrasilAPI."},
        )


@app.get("/api/company/{cnpj}/employees")
async def employees(cnpj: str):
    """Dispara o scraping do LinkedIn para os funcionarios da empresa."""
    company_name = ""
    company_legal = ""
    try:
        data = await brasilapi.fetch_company(cnpj)
        company_name = data.get("nome_fantasia") or data.get("razao_social") or ""
        company_legal = data.get("razao_social") or ""
    except Exception:
        # Se a BrasilAPI falhar, o scraper trata nomes vazios graciosamente.
        pass

    result = await linkedin_scraper.scrape_employees(
        brasilapi.only_digits(cnpj), company_name, company_legal
    )

    # Enriquece cada funcionario com o CPF, estilo Datastone:
    #   JBR (nome -> candidatos) + desambiguacao por empresa + cargo + CIDADE DA PESSOA.
    # A cidade e o cargo vem do proprio Bright (perfil do LinkedIn).
    if _cpf_ready() and isinstance(result.get("employees"), list):
        for emp in result["employees"]:
            nome = emp.get("name", "")
            try:
                cands = cpf_lookup.by_name(nome, limit=50)
            except Exception:
                continue

            if len(cands) == 1:
                emp["cpf_status"] = "resolved"
                emp["cpf"] = cands[0]["cpf"]
                continue

            # Varios homonimos: desambigua com sinais do Bright, se a Mk estiver ligada.
            if mkbuscas.enabled() and cands:
                res = await mkbuscas.disambiguate_multi(
                    cands,
                    company=company_name,
                    role=emp.get("title", ""),
                    city=emp.get("city", ""),
                )
                emp["cpf_status"] = res.get("status")
                if res.get("status") == "resolved":
                    emp["cpf"] = res.get("cpf")
                    tels = (res.get("pessoa") or {}).get("phones_mk", [])
                    if tels:
                        emp["phones_mk"] = tels
            else:
                emp["cpf_status"] = "ambiguous" if cands else "not_found"
                emp["cpf_candidates"] = len(cands)
    return result


async def _phones_for_cpf(cpf: str, modo: str = "celular", max_tel: int = 3,
                          fonte: str = "mk", modelo_id: str = "", modelo_nome: str = "",
                          cnpj: str = "") -> list[dict]:
    """Telefones para um CPF, filtrados/priorizados (celular atual primeiro).

    fonte: 'mk' (WorkAPI) | 'assertiva' (Localize). Ambos passam pelo mesmo
    refine_phones, garantindo o mesmo filtro (celular atual, dedupe, etc.).
    modelo_id/modelo_nome: se a fonte for assertiva, cada chamada é logada
    (custos.py) atribuída a esse modelo, pra rastrear gasto por planilha.
    """
    if not cpf:
        return []
    try:
        if fonte == "assertiva":
            if not assertiva.enabled():
                return []
            r = await assertiva.telefones_documento(cpf, tipo="CPF")
            custos.log_assertiva(modelo_id=modelo_id, modelo_nome=modelo_nome, cnpj=cnpj, cpf=cpf)
            if r.get("status") == "ok":
                return mkbuscas.refine_phones(r.get("telefones") or [], modo=modo, max_n=max_tel)
            return []
        if not mkbuscas.enabled():
            return []
        mk = await mkbuscas.consulta_cpf(cpf)
        if mk.get("status") == "ok":
            raw = mkbuscas._extract_phones(mk.get("data") or {})
            return mkbuscas.refine_phones(raw, modo=modo, max_n=max_tel)
    except Exception:
        pass
    return []


def _fmt_phone_digits(p: dict) -> str:
    ddd = str(p.get("ddd") or "").strip()
    num = str(p.get("number") or p.get("telefone") or "").strip()
    num = brasilapi.only_digits(num)
    if ddd and not num.startswith(ddd):
        return ddd + num
    return num or brasilapi.only_digits(str(p.get("telefone") or ""))


@app.post("/api/companies/search")
async def companies_search(payload: dict = Body(default={})):
    """Busca avancada de empresas (estilo Datastone) via Casa dos Dados.

    Body: {filtros: {...}, limite: int}. Retorna {status, total, empresas}.
    """
    filtros = payload.get("filtros") or {}
    limite = payload.get("limite") or 20
    offset = payload.get("offset") or 0
    # Base local (RFB) primeiro: sem cap de 20 e instantanea. Fallback: Casa dos Dados.
    if _cnpj_local():
        res = cnpj_lookup.search(filtros, limite=limite, offset=offset)
        res["fonte"] = "local"
        return res
    if not casadosdados.enabled():
        return {"status": "unavailable", "message": "Busca avancada indisponivel.", "empresas": []}
    res = await casadosdados.pesquisa_avancada(filtros, limite)
    res["fonte"] = "casadosdados"
    return res


@app.post("/api/prospeccao/pessoas")
async def prospeccao_pessoas(payload: dict = Body(default={})):
    """Busca de pessoas (perfil 'Clientes potenciais' estilo Datastone).

    PROXY com dados reais de HOJE: usa sócios da Receita (cargo = qualificação
    societária). Quando o dataset de perfis profissionais for ligado, este
    endpoint passa a incluir decisores não-sócios também.
    """
    if not _cnpj_local():
        return {"status": "unavailable", "message": "Base local indisponível.", "pessoas": []}
    filtros = payload.get("filtros") or {}
    limite = payload.get("limite") or 20
    offset = payload.get("offset") or 0
    return cnpj_lookup.search_pessoas(filtros, limite=limite, offset=offset)


def _resolve_modelo(modelo_id: str) -> str:
    """Nome de exibição do modelo, pra atribuir custo (custos.py)."""
    if not modelo_id:
        return ""
    if modelo_id == custos.CLIENTE_ID:
        return custos.CLIENTE_NOME
    for m in _modelos_load():
        if m.get("id") == modelo_id:
            return m.get("nome") or modelo_id
    return modelo_id


@app.get("/api/company/{cnpj}/leads")
async def company_leads(cnpj: str, decisores: bool = False,
                        modo_tel: str = "celular", max_tel: int = 3,
                        fonte_tel: str = "mk",
                        socios_modo: str = "todos", max_socios: int = 0,
                        modelo_id: str = ""):
    """Enriquece UMA empresa com contatos (socios do QSA + telefones).

    Retorna {status, empresa:{...}, contatos:[{tipo, nome, cargo, cpf, telefones[]}]}.
    modo_tel: 'celular' (padrao) | 'celular_fixo' | 'todos'.
    fonte_tel: 'mk' (WorkAPI) | 'assertiva' (Localize).
    max_tel: max de telefones por contato.
    socios_modo: 'todos' (padrao) | 'admin' (so socio-administrador/diretor/presidente).
    max_socios: 0 = sem limite; N = no maximo N socios (apos ordenar por qualificacao).
    modelo_id: modelo salvo (ou custos.CLIENTE_ID p/ planilha externa) — usado
    só pra atribuir o custo das consultas Assertiva desta montagem.
    Se decisores=true, tenta anexar decisores do LinkedIn (lento, best-effort).
    """
    # Base local (RFB) primeiro — instantânea e traz o QSA. Fallback: BrasilAPI.
    data = None
    if _cnpj_local():
        loc = cnpj_lookup.by_cnpj(cnpj)
        if loc.get("status") == "ok":
            data = loc["company"]
    if data is None:
        try:
            data = await brasilapi.fetch_company(cnpj)
        except LookupError:
            return {"status": "not_found", "cnpj": cnpj, "contatos": []}
        except Exception:
            return {"status": "error", "cnpj": cnpj, "message": "Falha BrasilAPI", "contatos": []}

    _enrich_qsa_cpf(data)
    qsa_all = [s for s in (data.get("qsa") or []) if isinstance(s, dict)]
    empresa = {
        "cnpj": brasilapi.format_cnpj(str(data.get("cnpj", "")) or cnpj),
        "razao_social": data.get("razao_social") or "",
        "nome_fantasia": data.get("nome_fantasia") or "",
        "municipio": data.get("municipio") or "",
        "uf": data.get("uf") or "",
        "porte": data.get("porte") or "",
        "capital_social": data.get("capital_social") or "",
        "cnae": data.get("cnae_fiscal_descricao") or "",
        "cnae_codigo": data.get("cnae_fiscal") or "",
        "situacao": data.get("descricao_situacao_cadastral") or "",
        "email": data.get("email") or "",
        "telefone_empresa": data.get("ddd_telefone_1") or "",
        # Campos extras p/ o export enriquecido (padrão da planilha modelo)
        "natureza_juridica": data.get("natureza_juridica") or "",
        "matriz_filial": data.get("matriz_filial") or "",
        "data_abertura": data.get("data_inicio_atividade") or "",
        "bairro": data.get("bairro") or "",
        "logradouro": data.get("logradouro") or "",
        "numero": data.get("numero") or "",
        "complemento": data.get("complemento") or "",
        "cep": data.get("cep") or "",
        "telefone_empresa_2": data.get("ddd_telefone_2") or "",
        "simples": data.get("opcao_simples") or "",
        "mei": data.get("opcao_mei") or "",
        "qtd_socios": len(qsa_all),
    }

    def _tel_payload(tels):
        return [{"raw": t.get("digits") or _fmt_phone_digits(t),
                 "display": t.get("telefone") or "", "categoria": t.get("categoria"),
                 "whatsapp": t.get("whatsapp")} for t in tels]

    socios = [s for s in (data.get("qsa") or []) if isinstance(s, dict)]
    # Preferir SÓCIO-ADMINISTRADOR como contato principal (Contato 1). A qualificação
    # societária (QSA) traz "Administrador"/"Sócio-Administrador"/"Diretor" — ordena
    # esses primeiro; mantém a ordem original como desempate.
    def _rank_socio(s):
        q = (s.get("qualificacao_socio") or s.get("qual") or "").lower()
        if "administrador" in q:
            return 0
        if "diretor" in q or "presidente" in q:
            return 1
        if "sócio" in q or "socio" in q:
            return 2
        return 3
    socios = sorted(socios, key=_rank_socio)
    if socios_modo == "admin":
        socios = [s for s in socios if _rank_socio(s) <= 1] or socios[:1]
    if max_socios and max_socios > 0:
        socios = socios[:max_socios]
    # Telefones de todos os socios em paralelo (era 1 a 1 => lento).
    modelo_nome = _resolve_modelo(modelo_id)
    tels_por_socio = await asyncio.gather(*[
        _phones_for_cpf(s.get("cpf_completo") or "", modo_tel, max_tel, fonte_tel,
                        modelo_id=modelo_id, modelo_nome=modelo_nome, cnpj=cnpj) for s in socios
    ])
    contatos = []
    for socio, tels in zip(socios, tels_por_socio):
        cpf = socio.get("cpf_completo") or ""
        contatos.append({
            "tipo": "socio",
            "nome": socio.get("nome_socio") or socio.get("nome") or "",
            "cargo": socio.get("qualificacao_socio") or socio.get("qual") or "",
            "cpf": cpf,
            "cpf_status": socio.get("cpf_status") or ("resolved" if cpf else "not_found"),
            "telefones": _tel_payload(tels),
        })

    if decisores and _cpf_ready():
        try:
            emp_res = await linkedin_scraper.scrape_employees(
                brasilapi.only_digits(cnpj), empresa["nome_fantasia"] or empresa["razao_social"],
                empresa["razao_social"],
            )
            for emp in (emp_res.get("employees") or [])[:10]:
                nome = emp.get("name", "")
                cpf = ""
                try:
                    cands = cpf_lookup.by_name(nome, limit=50)
                    if len(cands) == 1:
                        cpf = cands[0]["cpf"]
                except Exception:
                    pass
                tels = await _phones_for_cpf(cpf, modo_tel, max_tel, fonte_tel,
                                             modelo_id=modelo_id, modelo_nome=modelo_nome,
                                             cnpj=cnpj) if cpf else []
                contatos.append({
                    "tipo": "decisor",
                    "nome": nome,
                    "cargo": emp.get("title") or "",
                    "cpf": cpf,
                    "cpf_status": "resolved" if cpf else "ambiguous",
                    "telefones": [{"raw": t.get("digits") or _fmt_phone_digits(t),
                                   "display": t.get("telefone") or "", "categoria": t.get("categoria"),
                                   "whatsapp": t.get("whatsapp")} for t in tels],
                })
        except Exception:
            pass

    return {"status": "ok", "empresa": empresa, "contatos": contatos}


# Colunas de EMPRESA no export enriquecido (padrão da planilha modelo Datastone).
# (chave_no_payload | rótulo). Campos que ainda não temos dataset ficam em branco.
_EXP_EMPRESA_COLS = [
    ("cnpj", "CNPJ"),
    ("razao_social", "Razao Social"),
    ("nome_fantasia", "Nome Fantasia"),
    ("site", "Site (provavel)"),
    ("segmento", "Segmento"),
    ("setor_icp", "Setor ICP"),
    ("cnae_codigo", "CNAE Codigo"),
    ("cnae", "CNAE Descricao"),
    ("porte", "Porte"),
    ("funcionarios", "Funcionarios"),
    ("faturamento", "Faturamento"),
    ("matriz_filial", "Matriz/Filial"),
    ("natureza_juridica", "Natureza Juridica"),
    ("situacao", "Situacao Cadastral"),
    ("simples", "Simples Nacional"),
    ("data_abertura", "Data Abertura"),
    ("capital_social", "Capital Social"),
    ("uf", "UF"),
    ("municipio", "Municipio"),
    ("bairro", "Bairro"),
    ("logradouro", "Logradouro"),
    ("numero", "Numero"),
    ("cep", "CEP"),
    ("tel_empresa_1", "Tel Empresa 1"),
    ("tel_empresa_2", "Tel Empresa 2"),
    ("tel_empresa_3", "Tel Empresa 3"),
    ("email_empresa_1", "Email Empresa 1"),
    ("email_empresa_2", "Email Empresa 2"),
    ("email_empresa_3", "Email Empresa 3"),
    ("qtd_socios", "Qtd Socios"),
    ("qtd_membros", "Qtd Membros"),
    ("regional", "Regional"),
]

# Sufixos de cada bloco de contato (padrão "Contato N ...").
_EXP_CONTATO_FIELDS = [
    ("nome", "Nome"),
    ("cargo", "Cargo"),
    ("cpf", "CPF"),
    ("celular1", "Celular 1"),
    ("celular2", "Celular 2"),
    ("fixo", "Fixo"),
    ("whatsapp", "WhatsApp"),
    ("email1", "Email 1"),
    ("email2", "Email 2"),
]


def _split_contato(c: dict) -> dict:
    """Quebra um contato (com lista de telefones) nos campos do modelo:
    Celular 1/2, Fixo, WhatsApp, e-mails."""
    celulares, fixo, whats = [], "", False
    for t in (c.get("telefones") or []):
        disp = t.get("display") or t.get("raw") or ""
        cat = t.get("categoria") or ""
        if t.get("whatsapp"):
            whats = True
        if cat == "fixo":
            if not fixo:
                fixo = disp
        else:  # celular / celular_antigo
            celulares.append(disp)
    emails = c.get("emails") or []
    return {
        "nome": c.get("nome") or "",
        "cargo": c.get("cargo") or "",
        "cpf": c.get("cpf") or "",
        "celular1": celulares[0] if celulares else "",
        "celular2": celulares[1] if len(celulares) > 1 else "",
        "fixo": fixo,
        "whatsapp": "SIM" if whats else "",
        "email1": emails[0] if emails else "",
        "email2": emails[1] if len(emails) > 1 else "",
    }


def _site_from_email(email: str) -> str:
    email = (email or "").strip()
    if "@" in email:
        dom = email.split("@")[-1].strip().lower()
        generic = {"gmail.com", "hotmail.com", "outlook.com", "yahoo.com",
                   "yahoo.com.br", "uol.com.br", "bol.com.br", "terra.com.br",
                   "live.com", "icloud.com"}
        if dom and dom not in generic:
            return dom
    return ""


@app.post("/api/export/xlsx")
async def export_xlsx(payload: dict = Body(default={})):
    """Gera a planilha XLSX enriquecida (padrão Datastone), 1 linha por empresa.

    Body: {empresas: [{empresa:{...}, contatos:[{nome,cargo,cpf,telefones:[...],emails:[]}]}]}
    Compat: se vier {rows:[...]} (formato antigo por telefone), exporta assim mesmo.
    AutoFilter fica ligado em todas as colunas → dá pra adicionar mais filtros no Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    empresas = payload.get("empresas")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FF6A00")  # laranja do modelo
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospeccao"

    if empresas is not None:
        # layout: 'empresa' = 1 CNPJ por linha (blocos Contato 1..N lado a lado);
        #         'contato' = 1 contato por linha (empresa repetida a cada contato).
        layout = (payload.get("layout") or "empresa").strip().lower()

        def _emp_vals(emp: dict) -> list:
            vals = []
            for key, _ in _EXP_EMPRESA_COLS:
                if key == "site":
                    vals.append(_site_from_email(emp.get("email") or emp.get("email_empresa_1") or ""))
                elif key == "tel_empresa_1":
                    vals.append(emp.get("telefone_empresa") or "")
                elif key == "tel_empresa_2":
                    vals.append(emp.get("telefone_empresa_2") or "")
                elif key == "email_empresa_1":
                    vals.append(emp.get("email") or "")
                elif key == "simples":
                    v = emp.get("simples") or ""
                    vals.append("SIM" if str(v).upper() in ("S", "SIM", "1") else ("NAO" if v else ""))
                else:
                    vals.append(emp.get(key, ""))
            return vals

        norm = []
        max_contatos = 4  # como no modelo
        for item in empresas:
            emp = item.get("empresa") or {}
            contatos = [_split_contato(c) for c in (item.get("contatos") or [])]
            max_contatos = max(max_contatos, len(contatos))
            norm.append((emp, contatos))

        r_idx = 2
        if layout == "contato":
            # Cabeçalho: colunas de empresa + UM bloco de contato (sem prefixo "Contato N").
            headers = [lbl for _, lbl in _EXP_EMPRESA_COLS] + \
                      [f"Contato {lbl}" for _, lbl in _EXP_CONTATO_FIELDS]
            for emp, contatos in norm:
                base = _emp_vals(emp)
                # Empresa sem contato ainda gera 1 linha (só dados da empresa).
                for c in (contatos or [None]):
                    vals = list(base)
                    for fkey, _ in _EXP_CONTATO_FIELDS:
                        vals.append((c or {}).get(fkey, ""))
                    for col_idx, v in enumerate(vals, start=1):
                        ws.cell(row=r_idx, column=col_idx, value=v)
                    r_idx += 1
        else:
            # layout 'empresa': blocos Contato 1..N lado a lado.
            headers = [lbl for _, lbl in _EXP_EMPRESA_COLS]
            for n in range(1, max_contatos + 1):
                headers += [f"Contato {n} {lbl}" for _, lbl in _EXP_CONTATO_FIELDS]
            for emp, contatos in norm:
                vals = _emp_vals(emp)
                for c in contatos:
                    for fkey, _ in _EXP_CONTATO_FIELDS:
                        vals.append(c.get(fkey, ""))
                for col_idx, v in enumerate(vals, start=1):
                    ws.cell(row=r_idx, column=col_idx, value=v)
                r_idx += 1

        for col_idx, label in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        ws.freeze_panes = "C2"  # trava cabeçalho + CNPJ/Razão (como no modelo)
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{max(r_idx - 1, 1)}"
    else:
        # Fallback: formato antigo (uma linha por telefone).
        rows = payload.get("rows") or []
        legacy = [
            ("razao_social", "Razao Social"), ("nome_fantasia", "Nome Fantasia"),
            ("cnpj", "CNPJ"), ("municipio", "Municipio"), ("uf", "UF"),
            ("porte", "Porte"), ("cnae", "Atividade (CNAE)"), ("situacao", "Situacao"),
            ("contato_tipo", "Tipo Contato"), ("contato_nome", "Nome Contato"),
            ("contato_cargo", "Cargo"), ("contato_cpf", "CPF"), ("telefone", "Telefone"),
            ("tel_categoria", "Tipo Telefone"), ("validado", "Validado (telefone reverso)"),
            ("nome_donodozap", "Nome / Vinculo"),
        ]
        for col_idx, (_key, label) in enumerate(legacy, start=1):
            c = ws.cell(row=1, column=col_idx, value=label)
            c.font = header_font
            c.fill = header_fill
        for r_idx, row in enumerate(rows, start=2):
            for col_idx, (key, _label) in enumerate(legacy, start=1):
                ws.cell(row=r_idx, column=col_idx, value=row.get(key, ""))
        for i in range(1, len(legacy) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(legacy))}{max(len(rows) + 1, 1)}"

    # Aba "Fontes": de onde saem os dados (nome da API/base) — transparência.
    fonte_tel = (payload.get("fonte_tel") or "assertiva").lower()
    tel_fonte = "Assertiva Localize (API)" if fonte_tel == "assertiva" else "Mk Buscas / WorkAPI (API)"
    fontes = [
        ("Dados da empresa (razão, fantasia, CNAE, porte, situação, capital, endereço, matriz/filial, Simples/MEI)",
         "Receita Federal — Cadastro Nacional de CNPJ (base local RFB dos Dados Abertos)"),
        ("Telefone e e-mail da empresa", "Receita Federal (RFB) — quando disponível no cadastro"),
        ("Nome e cargo do contato/sócio", "Receita Federal — Quadro de Sócios (QSA)"),
        ("CPF do contato", "Base JBR (resolução nome + máscara do CPF)"),
        ("Telefones do contato (Celular/Fixo/WhatsApp)", tel_fonte),
        ("E-mail do contato", "Assertiva Localize (API)"),
        ("Verificação de telefone (pertence ao CPF)", "integralX / intelgrax-tel (WorkAPI) — telefone reverso"),
    ]
    wf = wb.create_sheet("Fontes")
    wf.cell(row=1, column=1, value="Campo").font = header_font
    wf.cell(row=1, column=2, value="Origem dos dados").font = header_font
    wf.cell(row=1, column=1).fill = header_fill
    wf.cell(row=1, column=2).fill = header_fill
    for i, (campo, origem) in enumerate(fontes, start=2):
        wf.cell(row=i, column=1, value=campo)
        wf.cell(row=i, column=2, value=origem)
    wf.column_dimensions["A"].width = 62
    wf.column_dimensions["B"].width = 66

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="capiblu-prospeccao.xlsx"'},
    )


# ============================================================
#  PROSPECÇÃO B2B — "UPLOAD MODELO"
#  Sobe uma planilha-modelo → detecta cabeçalhos e qual fonte/API preenche cada um
#  → gera a lista seguindo EXATAMENTE os cabeçalhos do modelo.
# ============================================================
import unicodedata as _ud


def _norm_hdr(h) -> str:
    s = _ud.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode("ascii").lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# campo | label | fonte | aliases (normalizados). Ordem importa p/ desempate.
_MODELO_CAMPOS = [
    ("cnpj", "CNPJ", "Receita Federal", ["cnpj"]),
    ("razao_social", "Razão Social", "Receita Federal",
     ["razao social", "razao", "empresa", "nome empresa", "razaosocial", "company", "nome"]),
    ("nome_fantasia", "Nome Fantasia", "Receita Federal", ["nome fantasia", "fantasia"]),
    ("cnae", "CNAE (descrição)", "Receita Federal",
     ["cnae", "atividade", "ramo", "segmento", "industria", "descricao cnae"]),
    ("cnae_codigo", "CNAE (código)", "Receita Federal", ["cnae codigo", "codigo cnae"]),
    ("porte", "Porte", "Receita Federal", ["porte", "tamanho", "funcionarios", "qtd funcionarios"]),
    ("situacao", "Situação Cadastral", "Receita Federal", ["situacao", "situacao cadastral", "status"]),
    ("natureza_juridica", "Natureza Jurídica", "Receita Federal",
     ["natureza", "natureza juridica", "tipo empresa"]),
    ("capital_social", "Capital Social", "Receita Federal", ["capital", "capital social", "faturamento"]),
    ("data_abertura", "Data de Abertura", "Receita Federal",
     ["data abertura", "abertura", "fundacao", "fundada", "data fundacao", "dt abertura", "dt inclusao"]),
    ("matriz_filial", "Matriz/Filial", "Receita Federal", ["matriz", "filial", "matriz filial"]),
    ("uf", "UF", "Receita Federal", ["uf", "estado"]),
    ("municipio", "Município", "Receita Federal", ["municipio", "cidade", "municipio2"]),
    ("bairro", "Bairro", "Receita Federal", ["bairro"]),
    ("logradouro", "Logradouro", "Receita Federal", ["logradouro", "rua", "endereco"]),
    ("numero", "Número", "Receita Federal", ["numero"]),
    ("cep", "CEP", "Receita Federal", ["cep"]),
    ("tel_empresa", "Telefone da Empresa", "Receita/Assertiva",
     ["telefone empresa", "tel empresa", "telefone da empresa", "fone empresa"]),
    ("email_empresa", "E-mail da Empresa", "Receita/Assertiva",
     ["email empresa", "e mail empresa", "email da empresa"]),
    ("qtd_socios", "Qtd Sócios", "Receita Federal", ["qtd socios", "numero de socios", "socios"]),
    ("contato_nome", "Nome do Contato", "Receita (QSA)",
     ["contato", "nome contato", "socio", "decisor", "responsavel", "representante", "nome socio"]),
    ("contato_cargo", "Cargo do Contato", "Receita (QSA)", ["cargo", "funcao", "qualificacao"]),
    ("contato_cpf", "CPF do Contato", "Base JBR", ["cpf", "cpf contato", "cpf socio"]),
    ("contato_celular2", "Celular 2 do Contato", "Assertiva/Mk",
     ["celular 2", "telefone 2", "celular2", "tel 2", "segundo telefone"]),
    ("contato_celular", "Celular do Contato", "Assertiva/Mk",
     ["celular", "celular 1", "telefone", "telefone 1", "movel", "tel", "fone", "telefone celular", "whatsapp"]),
    ("contato_fixo", "Fixo do Contato", "Assertiva/Mk", ["fixo", "telefone fixo"]),
    ("contato_whatsapp", "Tem WhatsApp?", "Assertiva/Mk",
     ["whatsapp", "possui whatsapp", "tem whatsapp", "zap", "wpp"]),
    ("contato_email", "E-mail do Contato", "Assertiva",
     ["email", "e mail", "email 1", "email contato", "e-mail"]),
]
_MODELO_LABEL = {k: lbl for k, lbl, _, _ in _MODELO_CAMPOS}
_MODELO_FONTE = {k: f for k, _, f, _ in _MODELO_CAMPOS}


def _match_header(header: str) -> Optional[str]:
    """Mapeia um cabeçalho do modelo para um campo conhecido (melhor correspondência)."""
    h = _norm_hdr(header)
    if not h:
        return None
    words = set(h.split())
    best, best_score = None, 0
    for campo, _lbl, _f, aliases in _MODELO_CAMPOS:
        for a in aliases:
            aw = a.split()
            score = 0
            if h == a:
                score = 100 + len(a)                      # match exato — melhor
            elif h.startswith(a + " "):
                score = 75 + len(a)                       # começa com o alias (palavra)
            elif all(w in words for w in aw):
                score = 55 + len(a)                       # todas as palavras do alias presentes
            elif h.endswith(" " + a):
                score = 40 + len(a)                       # termina com o alias
            elif a in h or h in a:
                score = 20 + len(a)                       # substring solta (fraco)
            if score > best_score:
                best, best_score = campo, score
    return best


def _extrai_modelo(lead: dict, campo: str, idx: int = 1):
    """Extrai o valor de um campo a partir de um lead {empresa, contatos}."""
    emp = lead.get("empresa") or {}
    emp_map = {
        "cnpj": emp.get("cnpj"), "razao_social": emp.get("razao_social"),
        "nome_fantasia": emp.get("nome_fantasia"), "cnae": emp.get("cnae"),
        "cnae_codigo": emp.get("cnae_codigo"), "porte": emp.get("porte"),
        "situacao": emp.get("situacao"), "natureza_juridica": emp.get("natureza_juridica"),
        "capital_social": emp.get("capital_social"), "data_abertura": emp.get("data_abertura"),
        "matriz_filial": emp.get("matriz_filial"), "uf": emp.get("uf"),
        "municipio": emp.get("municipio"), "bairro": emp.get("bairro"),
        "logradouro": emp.get("logradouro"), "numero": emp.get("numero"),
        "cep": emp.get("cep"), "tel_empresa": emp.get("telefone_empresa"),
        "email_empresa": emp.get("email"), "qtd_socios": emp.get("qtd_socios"),
    }
    if campo in emp_map:
        return emp_map[campo] if emp_map[campo] is not None else ""
    contatos = lead.get("contatos") or []
    if not contatos:
        return ""
    c = contatos[min(idx, len(contatos)) - 1] if idx >= 1 else contatos[0]
    split = _split_contato(c)
    cmap = {
        "contato_nome": split["nome"], "contato_cargo": split["cargo"],
        "contato_cpf": split["cpf"], "contato_celular": split["celular1"],
        "contato_celular2": split["celular2"], "contato_fixo": split["fixo"],
        "contato_whatsapp": split["whatsapp"], "contato_email": split["email1"],
    }
    return cmap.get(campo, "")


@app.post("/api/prospeccao/modelo/analisar")
async def modelo_analisar(file: UploadFile = File(...)):
    """Lê os cabeçalhos da planilha-modelo e diz qual campo/fonte preenche cada um."""
    try:
        content = await file.read()
        parsed, _aviso = sheet_reader.read_table(file.filename or "", content)
    except ValueError as exc:
        return {"status": "error", "message": str(exc)}
    except Exception as exc:
        return {"status": "error", "message": f"Falha ao ler modelo: {str(exc)[:150]}"}
    cols = parsed[0]["columns"] if parsed else []
    colunas = []
    for h in cols:
        # detecta índice de contato no cabeçalho (ex.: "Contato 2 Celular")
        m = re.search(r"contato\s*(\d+)", _norm_hdr(h))
        idx = int(m.group(1)) if m else 1
        campo = _match_header(h)
        colunas.append({
            "header": h, "campo": campo, "idx": idx,
            "campo_label": _MODELO_LABEL.get(campo) if campo else None,
            "fonte": _MODELO_FONTE.get(campo) if campo else None,
            "fillable": bool(campo),
        })
    campos_disp = [{"campo": k, "label": lbl, "fonte": f} for k, lbl, f, _ in _MODELO_CAMPOS]
    return {"status": "ok", "aba": parsed[0]["title"] if parsed else "",
            "colunas": colunas, "campos_disponiveis": campos_disp}


@app.post("/api/prospeccao/modelo/exportar")
async def modelo_exportar(payload: dict = Body(default={})):
    """Gera o XLSX seguindo os cabeçalhos do modelo.

    Body: {colunas:[{header, campo, idx}], empresas:[{empresa, contatos}]}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    colunas = payload.get("colunas") or []
    empresas = payload.get("empresas") or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo preenchido"
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1D4ED8")
    for i, col in enumerate(colunas, start=1):
        c = ws.cell(row=1, column=i, value=col.get("header"))
        c.font = hf
        c.fill = fill
    for r_idx, lead in enumerate(empresas, start=2):
        for i, col in enumerate(colunas, start=1):
            campo = col.get("campo")
            val = _extrai_modelo(lead, campo, int(col.get("idx") or 1)) if campo else ""
            if hasattr(val, "isoformat"):
                val = val.isoformat(sep=" ")[:19]
            ws.cell(row=r_idx, column=i, value=val)
    for i in range(1, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.freeze_panes = "A2"
    if colunas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{max(len(empresas) + 1, 1)}"
    # aba Fontes (só das colunas mapeadas)
    wf = wb.create_sheet("Fontes")
    wf.cell(row=1, column=1, value="Coluna do modelo").font = hf
    wf.cell(row=1, column=2, value="Campo").font = hf
    wf.cell(row=1, column=3, value="Origem").font = hf
    for j in (1, 2, 3):
        wf.cell(row=1, column=j).fill = fill
    for i, col in enumerate(colunas, start=2):
        campo = col.get("campo")
        wf.cell(row=i, column=1, value=col.get("header"))
        wf.cell(row=i, column=2, value=_MODELO_LABEL.get(campo, "— não preenchido") if campo else "— não preenchido")
        wf.cell(row=i, column=3, value=_MODELO_FONTE.get(campo, "") if campo else "")
    for col, w in (("A", 34), ("B", 26), ("C", 24)):
        wf.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="capiblu-modelo.xlsx"'})


# ---- Modelos SALVOS (persistidos em JSON no diretório de dados) ----
def _modelos_path() -> str:
    base = r"C:\capiblu_data" if os.path.isdir(r"C:\capiblu_data") else os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "capiblu_modelos.json")


def _modelos_load() -> list:
    import json as _json
    try:
        with open(_modelos_path(), encoding="utf-8") as fh:
            return _json.load(fh)
    except Exception:
        return []


@app.get("/api/custos/assertiva")
async def custos_assertiva(desde: str = "", ate: str = ""):
    """Resumo de custo Assertiva por modelo, no intervalo [desde, ate] (YYYY-MM-DD).

    Sem desde/ate: sem filtro de início/fim (mas o frontend manda os últimos 7 dias
    por padrão). 'ate' é inclusivo até o fim daquele dia.
    """
    import datetime as _dt
    desde_ts = None
    ate_ts = None
    try:
        if desde:
            desde_ts = _dt.datetime.strptime(desde, "%Y-%m-%d").timestamp()
        if ate:
            ate_ts = _dt.datetime.strptime(ate, "%Y-%m-%d").timestamp() + 86400 - 1
    except ValueError:
        return {"status": "error", "message": "Datas inválidas (use YYYY-MM-DD)."}
    res = custos.resumo(desde_ts, ate_ts)
    res["status"] = "ok"
    res["cliente_id"] = custos.CLIENTE_ID
    res["cliente_nome"] = custos.CLIENTE_NOME
    return res


def _modelos_save(lst: list) -> None:
    import json as _json
    with open(_modelos_path(), "w", encoding="utf-8") as fh:
        _json.dump(lst, fh, ensure_ascii=False, indent=1)


@app.get("/api/prospeccao/modelo/campos")
async def modelo_campos():
    """Catálogo de campos disponíveis (para o construtor de modelos)."""
    return {"campos": [{"campo": k, "label": lbl, "fonte": f} for k, lbl, f, _ in _MODELO_CAMPOS]}


@app.get("/api/prospeccao/modelos")
async def modelos_listar():
    return {"modelos": _modelos_load()}


@app.post("/api/prospeccao/modelos")
async def modelos_salvar(payload: dict = Body(default={})):
    """Cria/atualiza um modelo salvo. Body: {id?, nome, colunas:[{header,campo,idx}]}."""
    nome = (payload.get("nome") or "").strip()
    colunas = payload.get("colunas") or []
    if not nome:
        return {"status": "error", "message": "Informe um nome para o modelo."}
    if not colunas:
        return {"status": "error", "message": "O modelo precisa de ao menos uma coluna."}
    lst = _modelos_load()
    mid = payload.get("id")
    if mid:  # atualização
        for m in lst:
            if m.get("id") == mid:
                m["nome"], m["colunas"] = nome, colunas
                _modelos_save(lst)
                return {"status": "ok", "modelo": m}
    novo = {"id": uuid.uuid4().hex[:10], "nome": nome, "colunas": colunas}
    lst.append(novo)
    _modelos_save(lst)
    return {"status": "ok", "modelo": novo}


@app.delete("/api/prospeccao/modelos/{mid}")
async def modelos_excluir(mid: str):
    lst = [m for m in _modelos_load() if m.get("id") != mid]
    _modelos_save(lst)
    return {"status": "ok"}


# ============================================================
#  ENRIQUECER LISTA (upload XLSX -> qualifica + telefone + verifica)
# ============================================================

# Catálogo de campos que o usuário pode optar por adicionar.
# grupo | [(key, rótulo)]. Prefixo da key indica a fonte:
#   rfb_ = Receita local (grátis) · as_ = Assertiva · vf_ = verificação integralX
_ENRICH_CATALOG = [
    {"grupo": "Empresa (Receita Federal)", "fonte": "RFB local (instantâneo)", "campos": [
        ("rfb_razao", "Razão Social (RFB)"), ("rfb_fantasia", "Nome Fantasia"),
        ("rfb_situacao", "Situação Cadastral"), ("rfb_cnae_cod", "CNAE Código"),
        ("rfb_cnae", "CNAE Descrição"), ("rfb_porte", "Porte"),
        ("rfb_capital", "Capital Social"), ("rfb_natureza", "Natureza Jurídica"),
        ("rfb_abertura", "Data Abertura"), ("rfb_matriz", "Matriz/Filial"),
        ("rfb_endereco", "Endereço"), ("rfb_municipio", "Município (RFB)"),
        ("rfb_uf", "UF (RFB)"), ("rfb_cep", "CEP"),
        ("rfb_tel1", "Telefone Empresa 1"), ("rfb_tel2", "Telefone Empresa 2"),
        ("rfb_email", "E-mail Empresa (RFB)"), ("rfb_qtd_socios", "Qtd Sócios"),
        ("rfb_socio1", "Sócio Principal"), ("rfb_simples", "Simples"), ("rfb_mei", "MEI"),
    ]},
    {"grupo": "Empresa – Telefone (Assertiva)", "fonte": "Assertiva Localize (consumo)", "campos": [
        ("as_empresa_tel", "Telefone Empresa (Assertiva)"),
        ("as_empresa_tel2", "Telefone Empresa 2 (Assertiva)"),
        ("as_empresa_email", "E-mail Empresa (Assertiva)"),
        ("as_empresa_whatsapp", "WhatsApp Empresa"),
    ]},
    {"grupo": "Sócios – contato pessoal (Assertiva)", "fonte": "JBR (CPF) + Assertiva (consumo)", "campos": [
        ("so_socio1_nome", "Sócio 1 Nome"),
        ("so_socio1_cpf", "Sócio 1 CPF"),
        ("so_socio1_celular", "Sócio 1 Celular"),
        ("so_socio1_whatsapp", "Sócio 1 WhatsApp"),
        ("so_socio1_email", "Sócio 1 E-mail"),
        ("so_socio2_nome", "Sócio 2 Nome"),
        ("so_socio2_cpf", "Sócio 2 CPF"),
        ("so_socio2_celular", "Sócio 2 Celular"),
        ("so_todos_tel", "Todos os sócios (nome — celular)"),
    ]},
    {"grupo": "Verificação de telefone (integralX)", "fonte": "WorkAPI intelgrax-tel (consumo)", "campos": [
        ("vf_telefone", "Telefone Verificado"),
        ("vf_status", "Status Verificação"),
        ("vf_vinculos", "Nº de Vínculos"),
    ]},
]
_ENRICH_KEYS = {k for g in _ENRICH_CATALOG for (k, _) in g["campos"]}

# Store em memória das planilhas enviadas (ferramenta local, 1 usuário).
_UPLOADS: dict[str, dict] = {}


def _guess_cnpj_col(columns: list[str]) -> str:
    for c in columns:
        if "cnpj" in (c or "").strip().lower():
            return c
    return columns[0] if columns else ""


@app.get("/api/enrich/catalog")
async def enrich_catalog():
    return {"grupos": [
        {"grupo": g["grupo"], "fonte": g["fonte"],
         "campos": [{"key": k, "label": lbl} for k, lbl in g["campos"]]}
        for g in _ENRICH_CATALOG
    ], "assertiva_ok": assertiva.enabled(), "integralx_ok": bool(mkbuscas.TEL_AUTH_VALUE)}


@app.post("/api/enrich/upload")
async def enrich_upload(file: UploadFile = File(...)):
    """Recebe XLSX/XLS/CSV/TSV e devolve as abas, colunas e um preview (sem enriquecer).

    Parsing robusto (sheet_reader): encoding/delimitador auto, cabeçalho bagunçado,
    células com erro, floats de CNPJ, abas quebradas.
    """
    try:
        content = await file.read()
    except Exception as exc:
        return {"status": "error", "message": f"Falha ao receber arquivo: {str(exc)[:150]}"}
    if not content:
        return {"status": "error", "message": "Arquivo vazio."}
    try:
        parsed, aviso = sheet_reader.read_table(file.filename or "", content)
    except ValueError as exc:
        return {"status": "error", "message": str(exc)}
    except Exception as exc:
        return {"status": "error", "message": f"Não consegui ler a planilha: {str(exc)[:150]}"}

    up_id = uuid.uuid4().hex[:12]
    store, sheets = {}, []
    for sh in parsed:
        title = sh["title"]
        # evita colisão de nomes de aba
        base, n = title, 2
        while title in store:
            title = f"{base} ({n})"; n += 1
        store[title] = {"columns": sh["columns"], "rows": sh["rows"]}
        sheets.append({"name": title, "columns": sh["columns"], "linhas": len(sh["rows"]),
                       "preview": sh["rows"][:3], "cnpj_col": _guess_cnpj_col(sh["columns"])})
    _UPLOADS[up_id] = store
    return {"status": "ok", "upload_id": up_id, "sheets": sheets, "aviso": aviso}


async def _enrich_cnpj(cnpj: str, want: set) -> dict:
    """Enriquece um CNPJ com os campos pedidos. Só chama Assertiva/integralX se
    houver campos daquela fonte selecionados (controla consumo)."""
    out = {k: "" for k in want}
    # normaliza a célula (float/sci-notation) antes de extrair dígitos
    digits = brasilapi.only_digits(str(sheet_reader.normalize_cell(cnpj) or ""))
    need_rfb = any(k.startswith("rfb_") for k in want)
    need_as = any(k.startswith("as_") for k in want)
    need_so = any(k.startswith("so_") for k in want)
    need_vf = any(k.startswith("vf_") for k in want)
    # Excel costuma comer o zero à esquerda do CNPJ (guarda como número) — completa 14.
    if 8 <= len(digits) < 14:
        digits = digits.zfill(14)

    company = None
    if (need_rfb or need_as or need_vf or need_so) and _cnpj_local():
        loc = cnpj_lookup.by_cnpj(digits)
        if loc.get("status") == "ok":
            company = loc["company"]
    if company:
        qsa = company.get("qsa") or []
        s1 = qsa[0] if qsa else {}
        mp = {
            "rfb_razao": company.get("razao_social", ""),
            "rfb_fantasia": company.get("nome_fantasia", ""),
            "rfb_situacao": company.get("descricao_situacao_cadastral", ""),
            "rfb_cnae_cod": company.get("cnae_fiscal", ""),
            "rfb_cnae": company.get("cnae_fiscal_descricao", ""),
            "rfb_porte": company.get("porte", ""),
            "rfb_capital": company.get("capital_social", ""),
            "rfb_natureza": company.get("natureza_juridica", ""),
            "rfb_abertura": company.get("data_inicio_atividade", ""),
            "rfb_matriz": company.get("matriz_filial", ""),
            "rfb_endereco": " ".join(str(x) for x in [company.get("logradouro", ""),
                            company.get("numero", ""), company.get("bairro", "")] if x).strip(),
            "rfb_municipio": company.get("municipio", ""),
            "rfb_uf": company.get("uf", ""),
            "rfb_cep": company.get("cep", ""),
            "rfb_tel1": company.get("ddd_telefone_1", ""),
            "rfb_tel2": company.get("ddd_telefone_2", ""),
            "rfb_email": company.get("email", ""),
            "rfb_qtd_socios": len(qsa),
            "rfb_socio1": s1.get("nome_socio", ""),
            "rfb_simples": company.get("opcao_simples", ""),
            "rfb_mei": company.get("opcao_mei", ""),
        }
        for k in want:
            if k in mp:
                out[k] = mp[k]

    # Assertiva: telefone/email da empresa por CNPJ
    best_mobile = ""
    if (need_as or need_vf) and assertiva.enabled():
        try:
            r = await assertiva.telefones_documento(digits, tipo="CNPJ")
            if r.get("status") == "ok":
                tels = mkbuscas.refine_phones(r.get("telefones") or [], modo="todos", max_n=5)
                celus = [t for t in tels if t.get("categoria") == "celular"]
                fixos = [t for t in tels if t.get("categoria") == "fixo"]
                best_mobile = (celus[0]["digits"] if celus else (tels[0]["digits"] if tels else ""))
                if "as_empresa_tel" in want:
                    out["as_empresa_tel"] = celus[0]["digits"] if celus else (tels[0]["digits"] if tels else "")
                if "as_empresa_tel2" in want:
                    out["as_empresa_tel2"] = celus[1]["digits"] if len(celus) > 1 else (fixos[0]["digits"] if fixos else "")
                if "as_empresa_whatsapp" in want:
                    out["as_empresa_whatsapp"] = "SIM" if any(t.get("whatsapp") for t in tels) else ""
            if "as_empresa_email" in want:
                a = await assertiva.consulta_cnpj(digits)
                if a.get("status") == "ok":
                    emails = ((a.get("data") or {}).get("resposta") or {}).get("emails") or []
                    if emails:
                        e0 = emails[0]
                        out["as_empresa_email"] = e0.get("email") if isinstance(e0, dict) else str(e0)
        except Exception:
            pass

    # telefone/CPF do sócio p/ a verificação (preenchidos no bloco de sócios abaixo)
    socio_vphone, socio_vcpf = "", ""
    # Sócios — contato PESSOAL: resolve CPF do sócio (JBR) + celular/e-mail (Assertiva)
    if (need_so or need_vf) and company:
        _enrich_qsa_cpf(company)  # resolve cpf_completo de cada sócio via JBR
        socios = [s for s in (company.get("qsa") or []) if isinstance(s, dict)]
        resumo = []
        # processa no máx. 3 sócios (controla consumo)
        for idx, s in enumerate(socios[:3], start=1):
            nome = s.get("nome_socio") or ""
            cpf = s.get("cpf_completo") or ""
            cel, cel2, wa, email = "", "", "", ""
            if cpf and assertiva.enabled():
                try:
                    c = await assertiva.contato_cpf(cpf)
                    if c.get("status") == "ok":
                        tels = mkbuscas.refine_phones(c.get("telefones") or [], modo="celular_fixo", max_n=5)
                        celus = [t["digits"] for t in tels if t.get("categoria") == "celular"]
                        cel = celus[0] if celus else ""
                        cel2 = celus[1] if len(celus) > 1 else ""
                        wa = "SIM" if any(t.get("whatsapp") for t in tels) else ""
                        email = (c.get("emails") or [""])[0]
                except Exception:
                    pass
            # guarda o 1º sócio com celular como alvo da verificação
            if cel and not socio_vphone:
                socio_vphone, socio_vcpf = cel, cpf
                if not need_so:   # só precisávamos do telefone p/ verificar
                    break
            if idx == 1:
                for k, v in (("so_socio1_nome", nome), ("so_socio1_cpf", cpf),
                             ("so_socio1_celular", cel), ("so_socio1_whatsapp", wa),
                             ("so_socio1_email", email)):
                    if k in want:
                        out[k] = v
            elif idx == 2:
                for k, v in (("so_socio2_nome", nome), ("so_socio2_cpf", cpf),
                             ("so_socio2_celular", cel)):
                    if k in want:
                        out[k] = v
            if nome and cel:
                resumo.append(f"{nome}: {cel}")
        if "so_todos_tel" in want:
            out["so_todos_tel"] = " | ".join(resumo)

    # Verificação integralX: prioriza o telefone do SÓCIO e checa se PERTENCE ao
    # CPF dele (verificação real). Sem CPF/sócio, cai no telefone da empresa (contagem).
    if need_vf:
        phone = socio_vphone or best_mobile or brasilapi.only_digits(str(out.get("rfb_tel1") or ""))
        cpf_alvo = socio_vcpf if socio_vphone else ""
        if "vf_telefone" in want:
            out["vf_telefone"] = phone
        if phone and len(phone) >= 10 and mkbuscas.TEL_AUTH_VALUE:
            try:
                if cpf_alvo:
                    r = await mkbuscas.telefone_pertence(phone, cpf_alvo)
                    if r.get("status") == "no_access":
                        if "vf_status" in want:
                            out["vf_status"] = "sem acesso (chave integralX)"
                    elif r.get("status") == "ok":
                        if "vf_vinculos" in want:
                            out["vf_vinculos"] = r.get("total", 0)
                        if "vf_status" in want:
                            out["vf_status"] = ("pertence ao sócio" if r.get("atrelado")
                                                else ("compartilhado" if r.get("alerta_compartilhado")
                                                      else "não pertence"))
                    else:
                        if "vf_status" in want:
                            out["vf_status"] = "n/d"
                else:
                    rev = await mkbuscas.consulta_telefone(phone)
                    if rev.get("status") == "ok":
                        total = rev.get("total", 0)
                        if "vf_vinculos" in want:
                            out["vf_vinculos"] = total
                        if "vf_status" in want:
                            out["vf_status"] = ("compartilhado/lixo" if total >= 50
                                                else ("válido" if total >= 1 else "sem vínculo"))
                    elif rev.get("status") == "no_access":
                        if "vf_status" in want:
                            out["vf_status"] = "sem acesso (chave integralX)"
                    else:
                        if "vf_status" in want:
                            out["vf_status"] = "n/d"
            except Exception:
                if "vf_status" in want:
                    out["vf_status"] = "erro"
    return out


@app.post("/api/enrich/run")
async def enrich_run(payload: dict = Body(default={})):
    """Enriquece as linhas de uma aba. Body: {upload_id, sheet, cnpj_col, fields:[], limite}."""
    up_id = payload.get("upload_id")
    store = _UPLOADS.get(up_id)
    if not store:
        return {"status": "error", "message": "Upload expirado — reenvie a planilha."}
    sheet = payload.get("sheet") or next(iter(store))
    if sheet not in store:
        return {"status": "error", "message": "Aba não encontrada."}
    cnpj_col = payload.get("cnpj_col") or _guess_cnpj_col(store[sheet]["columns"])
    fields = [f for f in (payload.get("fields") or []) if f in _ENRICH_KEYS]
    if not fields:
        return {"status": "error", "message": "Selecione ao menos um campo para enriquecer."}
    limite = int(payload.get("limite") or 100)
    rows = store[sheet]["rows"][:limite]
    want = set(fields)

    # Concorrência controlada (Assertiva/integralX têm limite diário).
    sem = asyncio.Semaphore(6)

    async def _one(row):
        async with sem:
            enr = await _enrich_cnpj(row.get(cnpj_col, ""), want)
        merged = dict(row)
        merged.update(enr)
        return merged

    enriched = await asyncio.gather(*[_one(r) for r in rows])
    label_of = {k: lbl for g in _ENRICH_CATALOG for (k, lbl) in g["campos"]}
    added_cols = [{"key": f, "label": label_of.get(f, f)} for f in fields]
    return {"status": "ok", "sheet": sheet, "cnpj_col": cnpj_col,
            "base_cols": store[sheet]["columns"], "added_cols": added_cols,
            "rows": enriched, "enriquecidas": len(enriched),
            "total_aba": len(store[sheet]["rows"])}


@app.post("/api/enrich/export")
async def enrich_export(payload: dict = Body(default={})):
    """Gera XLSX com as colunas originais + as escolhidas. Body: {columns:[{key,label}], rows:[...]}."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = payload.get("columns") or []
    rows = payload.get("rows") or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista enriquecida"
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1D4ED8")
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=col.get("label", col.get("key")))
        c.font = hf
        c.fill = fill
    for r_idx, row in enumerate(rows, start=2):
        for i, col in enumerate(columns, start=1):
            v = row.get(col.get("key"), "")
            if hasattr(v, "isoformat"):
                v = v.isoformat(sep=" ")[:19]
            ws.cell(row=r_idx, column=i, value=v)
    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max(len(columns),1))}{max(len(rows)+1,1)}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="lista-enriquecida.xlsx"'})


# ---- Configurações de integração (admin) ----
# Admin é verificado pelo header X-User-Role, que SÓ o app-online (Render) envia
# após validar a sessão. O serviço de dados confia nele (só chega via proxy c/ segredo).

def _is_admin(request: Request) -> bool:
    return (request.headers.get("x-user-role") or "").lower() == "admin"


@app.post("/api/navlog")
async def navlog_registrar(request: Request, payload: dict = Body(default={})):
    """Log de navegação: qualquer usuário logado pode registrar (não é admin-only)."""
    email = request.headers.get("x-user-email") or ""
    navlog.registrar(
        email, payload.get("tab") or "",
        tipo=payload.get("tipo") or "", query=payload.get("query") or "", resultado=payload.get("resultado") or "",
    )
    return {"ok": True}


@app.get("/api/navlog/mine")
async def navlog_mine(request: Request, dias: int = 7):
    """Resumo de uso do próprio usuário (painel "Início") — não é admin-only."""
    email = request.headers.get("x-user-email") or ""
    res = navlog.resumo_usuario(email, dias)
    res["status"] = "ok"
    return res


@app.get("/api/navlog")
async def navlog_listar(request: Request, desde: str = "", ate: str = "", user: str = ""):
    """Histórico de navegação de todos os usuários — só admin."""
    if not _is_admin(request):
        return JSONResponse({"detail": "Requer admin."}, status_code=403)
    import datetime as _dt
    desde_ts = None
    ate_ts = None
    try:
        if desde:
            desde_ts = _dt.datetime.strptime(desde, "%Y-%m-%d").timestamp()
        if ate:
            ate_ts = _dt.datetime.strptime(ate, "%Y-%m-%d").timestamp() + 86400 - 1
    except ValueError:
        return {"status": "error", "message": "Datas inválidas (use YYYY-MM-DD)."}
    res = navlog.listar(desde_ts, ate_ts, user)
    res["status"] = "ok"
    return res


@app.get("/api/config")
async def config_get(request: Request):
    if not _is_admin(request):
        return JSONResponse({"detail": "Requer admin."}, status_code=403)
    tok = config_store.get("meetime_token") or os.environ.get("MEETIME_TOKEN", "")
    return {"meetime": {
        "configurado": bool(tok),
        "token_mascarado": (tok[:4] + "…" + tok[-4:]) if tok and len(tok) > 8 else ("definido" if tok else ""),
        "base_url": meetime._base_url(), "leads_path": meetime._leads_path(),
        "auth_header": meetime._auth_header(),
        "por_grupo": meetime.status_grupos(),
    }}


@app.post("/api/config/meetime")
async def config_meetime(request: Request, payload: dict = Body(default={})):
    if not _is_admin(request):
        return JSONResponse({"detail": "Requer admin."}, status_code=403)
    # Token específico de um grupo (multi-tenant) — não passa pelo config_store global.
    if payload.get("grupo_id"):
        meetime.set_token_grupo(payload["grupo_id"], str(payload.get("token") or "").strip())
        return {"status": "ok", "configurado": bool(meetime._token(payload["grupo_id"]))}
    updates = {}
    for k_in, k_cfg in (("token", "meetime_token"), ("base_url", "meetime_base_url"),
                        ("leads_path", "meetime_leads_path"), ("auth_header", "meetime_auth_header")):
        if payload.get(k_in) is not None:
            updates[k_cfg] = str(payload[k_in]).strip()
    config_store.set_many(updates)
    meetime._cache.clear()  # invalida cache (força rebaixar com a config nova)
    return {"status": "ok", "configurado": bool(meetime._token())}


# ---- Meetime: dedup (não prospectar quem já está no CRM) ----
# Cada usuário pertence a um grupo (X-User-Grupo, definido pelo admin em Usuários),
# e cada grupo tem seu próprio token/conta Meetime — usuários de grupos diferentes
# nunca cruzam dados de CRMs diferentes.

@app.get("/api/meetime/status")
async def meetime_status(request: Request, refresh: bool = False):
    grupo_id = request.headers.get("x-user-grupo") or ""
    if not meetime.enabled(grupo_id):
        return {"enabled": False}
    ex = await meetime.fetch_existing(force=refresh, grupo_id=grupo_id)
    return {"enabled": True, "status": ex.get("status"), "message": ex.get("message"),
            "total_cnpjs": len(ex.get("cnpjs") or []),
            "total_nomes": len(ex.get("nomes") or []), "cache": ex.get("cache")}


@app.post("/api/meetime/dedup")
async def meetime_dedup(request: Request, payload: dict = Body(default={})):
    """Body: {empresas:[{cnpj, razao_social}]}. Remove quem já está na Meetime
    (CNPJ exato OU nome por similaridade LIKE %). Retorna {novos, removidos}."""
    grupo_id = request.headers.get("x-user-grupo") or ""
    empresas = payload.get("empresas") or payload.get("candidatos") or []
    if not meetime.enabled(grupo_id):
        return {"status": "unavailable",
                "message": "Meetime não configurada para o seu grupo — peça ao admin para configurar em Usuários.",
                "novos": empresas, "removidos": []}
    ex = await meetime.fetch_existing(force=bool(payload.get("refresh")), grupo_id=grupo_id)
    if ex.get("status") and ex["status"] != "ok":
        return {"status": ex["status"], "message": ex.get("message"),
                "novos": empresas, "removidos": []}
    res = meetime.dedup(empresas, ex)
    res["status"] = "ok"
    return res


# ---- Serviço de dados: NÃO serve o frontend ----
# A tela de login e o app ficam no app-online (Render). Aqui é só endpoint de dados,
# acessível apenas pelo proxy com o segredo. Servir a UI aqui confundia (login dava
# "segredo de proxy"). O root só explica o que é.

@app.get("/")
async def index():
    return JSONResponse({
        "service": "capiblu-data",
        "mensagem": "Este é o serviço de DADOS interno do CapiBLU (uso via proxy). "
                    "Para acessar a plataforma, use o app online.",
        "app": "https://capiblu-app.onrender.com",
    })


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8010, reload=True)
