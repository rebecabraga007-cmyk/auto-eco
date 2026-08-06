"""Leitor robusto de planilhas para o "Enriquecer Lista".

Aceita XLSX, XLS (best-effort), CSV, TSV e texto delimitado, lidando com os
imprevistos comuns de planilha real:
- encoding variado (UTF-8/UTF-8-BOM/CP1252/Latin-1)
- delimitador variado no CSV (`,` `;` `\t` `|`) com autodetecção
- linhas vazias no topo antes do cabeçalho
- cabeçalhos em branco ou duplicados
- linhas irregulares (menos/mais colunas que o cabeçalho)
- células com erro de fórmula (#N/A, #REF!, etc.) e "nan"/"None"
- números com casa decimal (float) — ex.: CNPJ lido como 1963124000135.0
- notação científica (1.96312E+12)
- uma aba corrompida não derruba as demais

Saída uniforme: lista de abas [{"title", "columns":[str], "rows":[{col: val}]}].
"""
import csv
import io
import re

_MAX_ROWS = 200_000            # trava de segurança contra arquivo gigante
_ERR_TOKENS = {"", "#n/a", "#n/d", "#ref!", "#value!", "#div/0!", "#name?",
               "#null!", "#num!", "nan", "none", "null", "na"}


def normalize_cell(v):
    """Normaliza uma célula: float íntegro -> int, tira espaço, erros -> None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else v
    if isinstance(v, int):
        return v
    if isinstance(v, str):
        s = v.strip()
        if s.lower() in _ERR_TOKENS:
            return None
        # "1963124000135.0" -> "1963124000135" (float que virou texto)
        if re.fullmatch(r"-?\d+\.0+", s):
            return s.split(".")[0]
        # notação científica "1.963124E+12" -> inteiro, quando aplicável
        if re.fullmatch(r"-?\d+(\.\d+)?[eE][+-]?\d+", s):
            try:
                f = float(s)
                if f.is_integer():
                    return str(int(f))
            except ValueError:
                pass
        return s
    return v


def _normalize_headers(raw):
    headers, seen = [], {}
    for i, h in enumerate(raw):
        name = "" if h is None else str(h).strip()
        if not name:
            name = f"Coluna {i + 1}"
        key = name.lower()
        if key in seen:
            seen[key] += 1
            name = f"{name} ({seen[key]})"
        else:
            seen[key] = 1
        headers.append(name)
    return headers


def _table_from_matrix(matrix):
    """Acha o cabeçalho (1ª linha com conteúdo), monta dicts, tolera linhas irregulares."""
    # remove linhas totalmente vazias do topo
    start = 0
    while start < len(matrix) and not any(c not in (None, "") for c in matrix[start]):
        start += 1
    if start >= len(matrix):
        return [], []
    headers = _normalize_headers(matrix[start])
    ncol = len(headers)
    rows = []
    for raw in matrix[start + 1:]:
        cells = list(raw) + [None] * (ncol - len(raw))  # completa se vier curta
        # ignora linha inteiramente vazia
        if not any(c not in (None, "") for c in cells[:ncol]):
            continue
        rows.append({headers[i]: cells[i] for i in range(ncol)})
    return headers, rows


# --------------------------------------------------------------------------

def _read_xlsx(content: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        try:
            matrix = []
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i >= _MAX_ROWS:
                    break
                matrix.append([normalize_cell(c) for c in r])
            headers, rows = _table_from_matrix(matrix)
            if headers:
                sheets.append({"title": ws.title, "columns": headers, "rows": rows})
        except Exception:
            continue  # aba corrompida: pula, não derruba o arquivo
    return sheets


def _decode_text(content: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", errors="replace")


def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except Exception:
        # fallback: o delimitador mais frequente na 1ª linha não-vazia
        first = next((ln for ln in sample.splitlines() if ln.strip()), "")
        counts = {d: first.count(d) for d in ",;\t|"}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else ","


def _read_csv(content: bytes, title: str = "CSV"):
    text = _decode_text(content)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    sample = "\n".join(text.splitlines()[:20])
    delim = _sniff_delimiter(sample)
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    matrix = []
    for i, r in enumerate(reader):
        if i >= _MAX_ROWS:
            break
        matrix.append([normalize_cell(c) for c in r])
    headers, rows = _table_from_matrix(matrix)
    return [{"title": title, "columns": headers, "rows": rows}] if headers else []


def read_table(filename: str, content: bytes):
    """Dispatcher por extensão/conteúdo. Retorna (sheets, aviso|None).

    sheets: [{"title","columns","rows"}]. Levanta ValueError com msg amigável
    quando não consegue ler.
    """
    name = (filename or "").lower().strip()
    ext = name.rsplit(".", 1)[-1] if "." in name else ""
    aviso = None

    if ext in ("xlsx", "xlsm"):
        sheets = _read_xlsx(content)
    elif ext in ("csv", "tsv", "txt"):
        sheets = _read_csv(content, title=filename or "CSV")
    elif ext == "xls":
        # openpyxl não lê .xls antigo; tenta como XLSX (às vezes é .xlsx renomeado)
        try:
            sheets = _read_xlsx(content)
        except Exception:
            raise ValueError("Formato .xls antigo não suportado — salve como .xlsx ou .csv.")
    else:
        # sem extensão confiável: tenta XLSX, depois CSV
        try:
            sheets = _read_xlsx(content)
        except Exception:
            sheets = _read_csv(content, title=filename or "arquivo")
            aviso = "Extensão desconhecida — interpretado como texto delimitado (CSV)."

    if not sheets:
        raise ValueError("Nenhuma aba/linha legível encontrada na planilha.")
    return sheets, aviso
